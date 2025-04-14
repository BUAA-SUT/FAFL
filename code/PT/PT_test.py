import shutil
from typing import Dict
from PT import *
from openpyxl import load_workbook
from publicFun import *
import json
import sys
import os
import time
import dill
import zipfile
import csv
import subprocess
import threading
import multiprocessing
# import chardet
import codecs
import io
import logging
random.seed(1)

multiprocessing.set_start_method("fork", force=True)
multiprocessing.reduction.ForkingPickler = dill.Pickler


def copyFile(fileDir, tarDir, picknumber):
    pathDir = os.listdir(fileDir)  # 取图片的原始路径
    sample = random.sample(pathDir, picknumber)  # 随机选取picknumber数量的样本图片
    for i in range(len(sample)):
        shutil.copy(fileDir + sample[i], tarDir + "input{}.txt".format(i))
    return


def movefile(tarpath, tardir, oridir):
    if os.path.exists(tardir):
        if os.path.exists(tardir + '/' + tarpath):
            # os.remove(tarpath)
            pass
        else:
            shutil.move(oridir + tarpath, tardir)
    else:
        os.makedirs(tardir)
        shutil.move(oridir + tarpath, tardir)


def zip_file(src_dir):

    zip_name = src_dir + '.zip'

    z = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)

    for dirpath, dirnames, filenames in os.walk(src_dir):

        fpath = dirpath.replace(src_dir, '')

        fpath = fpath and fpath + os.sep or ''

        for filename in filenames:

            z.write(os.path.join(dirpath, filename), fpath+filename)

    z.close()


def un_zip(file_name):
    """unzip zip file"""
    zip_file = zipfile.ZipFile(file_name)
    tardir = '/Applications/work/data/MT/FS/PT/'
    if os.path.exists(file_name[:-4]):
        return
    for names in zip_file.namelist():
        zip_file.extract(names, tardir)
    zip_file.close()


def getMG(mu, mr_list, num_of_samples, originaloutput, ts):
    outputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT/RandomOutput{}.csv'.format(mu)
    csvFile = open(outputdir, "r")
    reader = csv.reader(csvFile)
    mutateoutput = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            mutateoutput[item[0]] = item[1].split("\n")
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    MGS = []
    for i in range(num_of_samples):
        MGs = []
        MG = [0] * len(mr_list)
        ts.setInputOutput("input{}".format(i), "output{}".format(i))
        original_output = MR().getResults(ts, mutateoutput)
        followup_ts = copy.copy(ts)
        for j in range(len(mr_list)):
            followup_ts.setInputOutput("input{}_{}".format(i, j), "output{}_{}".format(i, j))
            mr = mr_list[j]
            followup_output = mr.getResults(followup_ts, mutateoutput)
            expected_output = mr.getExpectedOutput(ts, original_output)
            isViolate = mr.assertViolation(expected_output, followup_output)
            if isViolate:
                MG[j] = 1
            elif (mutateoutput["output{}".format(i)] != originaloutput["output{}".format(i)] or
                  mutateoutput["output{}_{}".format(i, j)] != originaloutput["output{}_{}".format(i, j)]):
                MG[j] = 3
        MGs.append(MG)
        for m in range(len(mr_list)):
            MG = [0] * len(mr_list)
            ts.setInputOutput("input{}_{}".format(i, m), "output{}_{}".format(i, m))
            original_output = MR().getResults(ts, mutateoutput)
            followup_ts = copy.copy(ts)
            for n in range(len(mr_list)):
                followup_ts.setInputOutput("input{}_{}_{}".format(i, m, n), "output{}_{}_{}".format(i, m, n))
                mr = mr_list[n]
                followup_output = mr.getResults(followup_ts, mutateoutput)
                expected_output = mr.getExpectedOutput(ts, original_output)
                isViolate = mr.assertViolation(expected_output, followup_output)
                if isViolate:
                    MG[n] = 1
                elif (mutateoutput["output{}_{}".format(i, m)] != originaloutput["output{}_{}".format(i, m)] or
                      mutateoutput["output{}_{}_{}".format(i, m, n)] != originaloutput["output{}_{}_{}".format(i, m, n)]):
                    MG[n] = 3
            MGs.append(MG)
        MGS.append(MGs)
    return MGS


def getpf(mu, ts, num_of_samples, mr_list, outputdata):
    # pf = []
    # for i in range(num_of_samples):
    #     result = [0] * (len(mr_list) * (len(mr_list) + 1) + 1)
    #     ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(0, i))
    #     original_output = MR().getResults(ts)
    #     program_ts = ts
    #     program_ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(mu, i))
    #     program_output = MR().getResults(program_ts)
    #     isViolate = MR().assertViolation(original_output, program_output)
    #     if isViolate:
    #         result[0] = 1
    #
    #     for j in range(len(mr_list)):
    #         ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(0, i, j))
    #         original_output = MR().getResults(ts)
    #         program_ts = ts
    #         program_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(mu, i, j))
    #         program_output = MR().getResults(program_ts)
    #         isViolate = MR().assertViolation(original_output, program_output)
    #         if isViolate:
    #             result[j + 1] = 1
    #
    #     for m in range(len(mr_list)):
    #         for n in range(len(mr_list)):
    #             ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n),
    #                                        "output{}_{}_{}_{}.txt".format(0, i, m, n))
    #             original_output = MR().getResults(ts)
    #             program_ts = ts
    #             program_ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n), "output{}_{}_{}_{}.txt".format(mu, i, m, n))
    #             program_output = MR().getResults(program_ts)
    #             isViolate = MR().assertViolation(original_output, program_output)
    #             if isViolate:
    #                 result[len(mr_list) + 1 + m * len(mr_list) + n] = 1
    #     pf.append(result)
    pf = []
    for i in range(num_of_samples):
        result1 = [0] * (len(mr_list) * (len(mr_list) + 1) + 1)  # t1,t2,...,t13
        ts.setInputOutput("input{}".format(i), "output{}_{}".format(0, i))
        original_output = MR().getResults(ts, outputdata)
        program_ts = copy.copy(ts)
        program_ts.setInputOutput("input{}".format(i), "output{}_{}".format(mu, i))
        program_output = MR().getResults(program_ts, outputdata)
        isViolate = MR().assertViolation(original_output, program_output)
        if isViolate:
            result1[0] = 1

        for j in range(len(mr_list)):
            ts.setInputOutput("input{}_{}".format(i, j), "output{}_{}_{}".format(0, i, j))
            original_output = MR().getResults(ts, outputdata)
            program_ts = copy.copy(ts)
            program_ts.setInputOutput("input{}_{}".format(i, j), "output{}_{}_{}".format(mu, i, j))
            program_output = MR().getResults(program_ts, outputdata)
            isViolate = MR().assertViolation(original_output, program_output)
            if isViolate:
                result1[j + 1] = 1

        for m in range(len(mr_list)):
            for n in range(len(mr_list)):
                ts.setInputOutput("input{}_{}_{}".format(i, m, n),
                                  "output{}_{}_{}_{}".format(0, i, m, n))
                original_output = MR().getResults(ts, outputdata)
                program_ts = copy.copy(ts)
                program_ts.setInputOutput("input{}_{}_{}".format(i, m, n), "output{}_{}_{}_{}".format(mu, i, m, n))
                program_output = MR().getResults(program_ts, outputdata)
                isViolate = MR().assertViolation(original_output, program_output)
                if isViolate:
                    result1[len(mr_list) + 1 + m * len(mr_list) + n] = 1

        # t14,t15,...
        result2 = []
        for j in range(len(mr_list)):
            r1 = []
            for k in range(len(mr_list)):
                r2 = []
                ts.setInputOutput("input{}_{}_{}".format(i, j, k),
                                  "output{}_{}_{}_{}".format(0, i, j, k))
                original_output = MR().getResults(ts, outputdata)
                program_ts = copy.copy(ts)
                program_ts.setInputOutput("input{}_{}_{}".format(i, j, k), "output{}_{}_{}_{}".format(mu, i, j, k))
                program_output = MR().getResults(program_ts, outputdata)
                isViolate = MR().assertViolation(original_output, program_output)
                if isViolate:
                    result = 1
                else:
                    result = 0
                r2.append(result)
                for m in range(len(mr_list)):
                    ts.setInputOutput("input{}_{}_{}_{}".format(i, j, k, m),
                                      "output{}_{}_{}_{}_{}".format(0, i, j, k, m))
                    original_output = MR().getResults(ts, outputdata)
                    program_ts = copy.copy(ts)
                    program_ts.setInputOutput("input{}_{}_{}_{}".format(i, j, k, m),
                                              "output{}_{}_{}_{}_{}".format(mu, i, j, k, m))
                    program_output = MR().getResults(program_ts, outputdata)
                    isViolate = MR().assertViolation(original_output, program_output)
                    if isViolate:
                        result = 1
                    else:
                        result = 0
                    r2.append(result)
                for m in range(len(mr_list)):
                    for n in range(len(mr_list)):
                        ts.setInputOutput("input{}_{}_{}_{}_{}".format(i, j, k, m, n),
                                          "output{}_{}_{}_{}_{}_{}".format(0, i, j, k, m, n))
                        original_output = MR().getResults(ts, outputdata)
                        program_ts = copy.copy(ts)
                        program_ts.setInputOutput("input{}_{}_{}_{}_{}".format(i, j, k, m, n),
                                                  "output{}_{}_{}_{}_{}_{}".format(mu, i, j, k, m, n))
                        program_output = MR().getResults(program_ts, outputdata)
                        isViolate = MR().assertViolation(original_output, program_output)
                        if isViolate:
                            result = 1
                        else:
                            result = 0
                        r2.append(result)
                r1.append(r2)
            result2.append(r1)
        pf.append([result1, result2])

    return pf


def get_input():
    inputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT/RandomInput.csv'
    # inputdir = '/home/rdx/data/MT/STVR/PT/RandomInput.csv'
    # 建立空字典
    csvFile = open(inputdir, "r")
    # reader = csv.reader(csvFile)
    reader = csv.reader((line.replace('\0', '') for line in csvFile))
    # 建立空字典
    inputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            inputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()
    return inputcase


def dataTrans(mr, num_of_samples):
    inputs_write = '/Users/rendaixu/OneDrive/data/MT/STVR/PT/RandomInput.csv'
    # outputs_write = '/Users/rendaixu/OneDrive/data/MT/FS/PT/RandomOutput'
    inputs_read = '/Applications/work/data/MT/MFT/PT/RandomInput'
    # outputs_read = '/Applications/work/data/MT/FS/PT/RandomOutput'
    fileHeader = ["name", "value"]
    inputdata = [fileHeader]

    for i in range(num_of_samples):
        f = open(inputs_read+'/input'+str(i)+'.txt')
        data = ['input{}'.format(i), f.read()]
        inputdata.append(data)
        for j in range(mr):
            f = open(inputs_read + '/input{}_{}.txt'.format(i, j))
            data = ['input{}_{}'.format(i, j), f.read()]
            inputdata.append(data)
            for k in range(mr):
                f = open(inputs_read + '/input{}_{}_{}.txt'.format(i, j, k))
                data = ['input{}_{}_{}'.format(i, j, k), f.read()]
                inputdata.append(data)
    # 将数据存下来
    with open(inputs_write, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(inputdata)


def verify(mr, num_of_samples):
    inputs_read = '/Applications/work/data/MT/MFT/PT/RandomOutput'
    fileHeader = ["name", "value"]
    inputdata = [fileHeader]
    for i in range(num_of_samples):
        f = open(inputs_read+'/output0_'+str(i)+'.txt')
        data = ['output0_{}'.format(i), f.read()]
        inputdata.append(data)
        for j in range(mr):
            f = open(inputs_read + '/output0_{}_{}.txt'.format(i, j))
            data = ['output0_{}_{}'.format(i, j), f.read()]
            inputdata.append(data)
            for k in range(mr):
                f = open(inputs_read + '/output0_{}_{}_{}.txt'.format(i, j, k))
                data = ['output0_{}_{}_{}'.format(i, j, k), f.read()]
                inputdata.append(data)
    inputcase = get_input()
    lst1 = list(inputcase.items())
    lst2 = list(inputcase.items())
    for i in range(len(lst1)):
        if not lst1[i][1] == lst2[i][1]:
            print(i)


# def getOutput(mu, inputcase):
#     fileHeader = ["name", "value"]
#     outputdata = [fileHeader]
#     command = './Mutants/printtokens_v{}/print_tokens'.format(mu)
#     lst = list(inputcase.items())
#     for i in range(len(lst)):
#         # 将输入字符串写入临时文件
#         try:
#             output = subprocess.run(command, input=lst[i][1].encode(), stdout=subprocess.PIPE,
#                                     stderr=subprocess.STDOUT, shell=True).stdout.decode()
#         except UnicodeDecodeError:
#             print("Mutant{} 编码格式问题，跳过".format(mu))
#             return 0
#         data = ['output'+lst[i][0][5:], output]
#         outputdata.append(data)
#     outputs_write = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT/RandomOutput{}.csv'.format(mu)
#     # outputs_write = '/home/rdx/data/MT/STVR/PT/RandomOutput{}.csv'.format(mu)
#     with open(outputs_write, 'w', newline='') as csvfile:
#         writer = csv.writer(csvfile)
#         writer.writerows(outputdata)
#     print('第{}个已完成'.format(mu))


def mutaterate(mu):
    outputdir1 = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT/RandomOutput0.csv'
    outputdir2 = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT/RandomOutput{}.csv'.format(mu)
    csvFile = open(outputdir1, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase1 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase1[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    csvFile = open(outputdir2, "r")
    reader = csv.reader(csvFile)
    outputcase2 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase2[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    num = 0
    outputcase1 = list(outputcase1.values())
    outputcase2 = list(outputcase2.values())
    index = []
    for i in range(len(outputcase1)):
        if not outputcase1[i] == outputcase2[i]:
            num += 1
            index.append(i)
    rate = num / (len(outputcase1))
    print(mu, round(rate*100, 2))
    return index


def target_function(processor):
    processor.getOutput()


class MutantProcessor:
    def __init__(self, mu, inputcase):
        self.mu = mu
        self.inputcase = inputcase

    def getOutput(self):
        fileHeader = ["name", "value"]
        outputdata = [fileHeader]
        command = './Mutants/printtokens_v{}/print_tokens'.format(self.mu)
        lst = list(self.inputcase.items())

        for i in range(len(lst)):
            # 将输入字符串写入临时文件
            try:
                output = subprocess.run(
                    command,
                    input=lst[i][1].encode(),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    shell=True
                ).stdout.decode()
            except UnicodeDecodeError:
                print("Mutant{} 编码格式问题，跳过".format(self.mu))
                return 0

            data = ['output' + lst[i][0][5:], output]
            outputdata.append(data)

        outputs_write = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT/RandomOutput{}.csv'.format(self.mu)

        with open(outputs_write, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(outputdata)

        print('第{}个已完成'.format(self.mu))


def run_getOutput(i, inputcase, timeout):
    # 创建 MutantProcessor 类的实例
    processor = MutantProcessor(mu=i, inputcase=inputcase)

    # 创建 multiprocessing.Process 实例
    p = multiprocessing.Process(target=target_function, args=(processor,))
    p.start()

    # 等待进程完成，设置超时
    p.join(timeout)

    # 如果进程超时，终止进程
    if p.is_alive():
        p.terminate()
        p.join()  # 确保进程结束
        print("Mutant{} 超时，跳过".format(i))
    else:
        print("Mutant{} 完成".format(i))

    # p = multiprocessing.Process(target=getOutput, args=(i, inputcase))
    # try:
    #     p.start()
    #     p.join(timeout)
    #
    #     if p.is_alive():
    #         p.terminate()
    #         p.join()
    #         logging.warning(f"Mutant{i} 超时，已跳过")
    #     else:
    #         logging.info(f"Mutant{i} 正常结束")
    #
    # except Exception as e:
    #     logging.error(f"Mutant{i} 处理时出错: {e}")
    # finally:
    #     if p.is_alive():
    #         p.terminate()
    #         p.join()


def rank_of_element(lst, element):
    sorted_list = sorted(lst, reverse=True)
    return sorted_list.index(element) + 1


if __name__ == "__main__":
    num_case = 100
    num_mr = 11
    # mutant = [2, 32, 58, 88, 98, 100, 110, 113, 120, 124, 134, 135, 141, 143, 147, 156, 161, 165,
    #           167, 174, 192, 193, 201, 202, 205, 210, 215, 217, 220, 240, 260, 284, 333, 351, 352, 353]
    mutant = [2, 32, 58, 88, 120, 134, 161, 201, 202, 205, 210, 215,
              220, 240, 260, 354, 355, 360, 364, 366, 368]
    # mutant = [351, 352, 353]
    # dataTrans(mr, num_of_samples)
    # inputcase = get_input()
    # timeout = 600  # 设定阈值为**秒
    # # run_getOutput(0, inputcase, timeout)
    # for i in mutant:
    #     run_getOutput(i, inputcase, timeout)
    # for mu in range(num_mu+1):
    #     index = mutaterate(mu)
    #
    # outputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT/RandomOutput0.csv'
    # csvFile = open(outputdir, "r")
    # reader = csv.reader(csvFile)
    # # 建立空字典
    # originaloutput = {}
    # for item in reader:
    #     # 忽略第一行
    #     if reader.line_num == 1:
    #         continue
    #     try:
    #         originaloutput[item[0]] = item[1].split("\n")
    #     except:
    #         print(item)
    #         print(reader.line_num)
    #         break
    # csvFile.close()
    # mr_list = [MR1(), MR2(), MR3(), MR4(), MR5(), MR6(), MR7(), MR8(), MR9(), MR10(), MR11()]
    # ts = TestCase()
    # # MGS_set = []
    # # SMGS_set = []
    # # for mu in range(1, 24):
    # #     MGS, SMGS = getMG(mu, mr_list, num_of_samples, originaloutput, ts)
    # #     MGS_set.append(MGS)
    # #     SMGS_set.append(SMGS)
    #
    # # verify(mr, num_of_samples)
    # # # myenv = MyEnv()
    # # # myenv.CreateWorkingDirs()
    project = 'MSBF'
    row = 1
    string = 'PT'
    path = '/Applications/work/data/MT/' + project + '/Result/result8_nofs.xlsx'  # '+sys.argv[1][:-1]+'
    # datadir = '/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statementResult.csv'
    wb = load_workbook(path)
    if string not in wb.sheetnames:
        ws = wb.create_sheet(string)
    del wb[string]
    ws = wb.create_sheet(string)
    # MG_set = []
    # MG_set2 = []
    # ExecutableS_set = []
    # Sus_set = []
    # Flag_set = []
    # # csvFile = open(datadir, "r")
    # # reader = csv.reader(csvFile)
    # # # 建立空字典
    # # originaldata = {}
    # # for item in reader:
    # #     originaldata[item[0]] = item[1:-17]
    # # csvFile.close()
    # # ExecutableS = originaldata.get('inputs')
    # # ExecutableS = [int(x) for x in ExecutableS]
    # # ExelinesS = []
    # # for i in range(num_case):
    # #     Exelines = []
    # #     candidate = originaldata.get('input{}'.format(i))
    # #     candidate = [int(x) for x in candidate]
    # #     indices = [i for i in range(len(candidate)) if candidate[i] == 1]
    # #     Exelines.append([ExecutableS[i] for i in indices])
    # #     for j in range(num_mr):
    # #         candidate = originaldata.get('input{}_{}'.format(i, j))
    # #         candidate = [int(x) for x in candidate]
    # #         indices = [i for i in range(len(candidate)) if candidate[i] == 1]
    # #         Exelines.append([ExecutableS[i] for i in indices])
    # #     for m in range(num_mr):
    # #         for n in range(num_mr):
    # #             candidate = originaldata.get('input{}_{}_{}'.format(i, m, n))
    # #             candidate = [int(x) for x in candidate]
    # #             indices = [i for i in range(len(candidate)) if candidate[i] == 1]
    # #             Exelines.append([ExecutableS[i] for i in indices])
    # #     ExelinesS.append(Exelines)
    # # data = {
    # #     'Exec': ExecutableS, 'Exel': ExelinesS
    # # }
    # # json_str = json.dumps(data)
    # # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/statements.json', 'w') as f:
    # #     json.dump(json_str, f)
    # Ind = []
    for mu in mutant:
        # datadir = '/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statementResult{}.csv'.format(mu)
        # csvFile = open(datadir, "r")
        # reader = csv.reader(csvFile)
        # # 建立空字典
        # originaldata = {}
        # result = {}
        # for item in reader:
        #     originaldata[item[0]] = item[1:-1]
        #     if reader.line_num == 1:
        #         continue
        #     result[item[0]] = item[-1:]
        # csvFile.close()
        # ExecutableS = originaldata.get('inputs')
        # ExecutableS = [int(x) for x in ExecutableS]
        # # ExelineS = []
        # Result = []
        # for i in range(num_case):
        #     Exelines = []
        #     r = []
        #     candidate = originaldata.get('input{}'.format(i))
        #     candidate = [int(x) for x in candidate]
        #     indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #     Exelines.append([ExecutableS[i] for i in indices])
        #     a = result.get('input{}'.format(i))
        #     a = [int(x) for x in a]
        #     r.append(a[0])
        #     for j in range(num_mr):
        #         candidate = originaldata.get('input{}_{}'.format(i, j))
        #         candidate = [int(x) for x in candidate]
        #         indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #         Exelines.append([ExecutableS[i] for i in indices])
        #         a = result.get('input{}_{}'.format(i, j))
        #         a = [int(x) for x in a]
        #         r.append(a[0])
        #     for m in range(num_mr):
        #         for n in range(num_mr):
        #             candidate = originaldata.get('input{}_{}_{}'.format(i, m, n))
        #             candidate = [int(x) for x in candidate]
        #             indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #             Exelines.append([ExecutableS[i] for i in indices])
        #             a = result.get('input{}_{}_{}'.format(i, m, n))
        #             a = [int(x) for x in a]
        #             r.append(a[0])
        #     Result.append(r)
        #     ExelineS.append(Exelines)
        # MGS = getMG(mu, mr_list, num_case, originaloutput, ts)
        # Flag = [0] * len(ExecutableS)
        # if mu == 2:
        #     Flag[ExecutableS.index(465)] = 1
        # # elif mu == 32:
        # #     Flag[ExecutableS.index(418)] = 1
        # # elif mu == 58:
        # #     Flag[ExecutableS.index(328)] = 1
        # # elif mu == 88:
        # #     Flag[ExecutableS.index(558)] = 1
        # # elif mu == 98:
        # #     Flag[ExecutableS.index(232)] = 1
        # # elif mu == 100:
        # #     Flag[ExecutableS.index(228)] = 1
        # # elif mu == 110:
        # #     Flag[ExecutableS.index(232)] = 1
        # # elif mu == 113:
        # #     Flag[ExecutableS.index(249)] = 1
        # # elif mu == 120:
        # #     Flag[ExecutableS.index(258)] = 1
        # # elif mu == 124:
        # #     Flag[ExecutableS.index(228)] = 1
        # # elif mu == 134:
        # #     Flag[ExecutableS.index(245)] = 1
        # # elif mu == 135:
        # #     Flag[ExecutableS.index(231)] = 1
        # # elif mu == 141:
        # #     Flag[ExecutableS.index(232)] = 1
        # # elif mu == 143:
        # #     Flag[ExecutableS.index(231)] = 1
        # # elif mu == 147:
        # #     Flag[ExecutableS.index(231)] = 1
        # # elif mu == 156:
        # #     Flag[ExecutableS.index(231)] = 1
        # # elif mu == 161:
        # #     Flag[ExecutableS.index(213)] = 1
        # # elif mu == 165:
        # #     Flag[ExecutableS.index(232)] = 1
        # # elif mu == 167:
        # #     Flag[ExecutableS.index(231)] = 1
        # # elif mu == 174:
        # #     Flag[ExecutableS.index(213)] = 1
        # # elif mu == 192:
        # #     Flag[ExecutableS.index(213)] = 1
        # # elif mu == 193:
        # #     Flag[ExecutableS.index(232)] = 1
        # elif mu == 201:
        #     Flag[ExecutableS.index(245)] = 1
        # elif mu == 202:
        #     Flag[ExecutableS.index(245)] = 1
        # elif mu == 205:
        #     Flag[ExecutableS.index(209)] = 1
        # elif mu == 210:
        #     Flag[ExecutableS.index(254)] = 1
        # elif mu == 215:
        #     Flag[ExecutableS.index(245)] = 1
        # elif mu == 217:
        #     Flag[ExecutableS.index(228)] = 1
        # elif mu == 220:
        #     Flag[ExecutableS.index(213)] = 1
        # elif mu == 240:
        #     Flag[ExecutableS.index(465)] = 1
        # elif mu == 260:
        #     Flag[ExecutableS.index(463)] = 1
        # elif mu == 284:
        #     Flag[ExecutableS.index(293)] = 1
        # elif mu == 333:
        #     Flag[ExecutableS.index(513)] = 1
        # elif mu == 351:
        #     Flag[ExecutableS.index(209)] = 1
        # elif mu == 352:
        #     Flag[ExecutableS.index(209)] = 1
        # elif mu == 353:
        #     Flag[ExecutableS.index(209)] = 1
        # elif mu == 354:
        #     Flag[ExecutableS.index(463)] = 1
        # elif mu == 355:
        #     Flag[ExecutableS.index(465)] = 1
        # elif mu == 356:
        #     Flag[ExecutableS.index(229)] = 1
        # elif mu == 357:
        #     Flag[ExecutableS.index(280)] = 1
        # elif mu == 358:
        #     Flag[ExecutableS.index(463)] = 1
        # elif mu == 359:
        #     Flag[ExecutableS.index(282)] = 1
        # elif mu == 360:
        #     Flag[ExecutableS.index(463)] = 1
        # elif mu == 361:
        #     Flag[ExecutableS.index(556)] = 1
        # elif mu == 362:
        #     Flag[ExecutableS.index(562)] = 1
        # elif mu == 363:
        #     Flag[ExecutableS.index(562)] = 1
        # elif mu == 364:
        #     Flag[ExecutableS.index(197)] = 1
        # elif mu == 365:
        #     Flag[ExecutableS.index(559)] = 1
        # elif mu == 366:
        #     Flag[ExecutableS.index(197)] = 1
        # elif mu == 367:
        #     Flag[ExecutableS.index(277)] = 1
        # elif mu == 368:
        #     Flag[ExecutableS.index(418)] = 1
        # elif mu == 369:
        #     Flag[ExecutableS.index(441)] = 1
        # elif mu == 370:
        #     Flag[ExecutableS.index(556)] = 1
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # MGS = data['MGS']
        # ExecutableS = data['Exec']
        # ExelineS = data['Exel']
        # Flag = data['Flag']
        # sus, metric = Sus(MGS, ExecutableS, ExelineS)
        # # AllF = data['AllF']
        # FAsus, FAmetric, percent = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # data = {
        #     'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
        #     'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
        #     'Exec': ExecutableS, 'Exel': ExelineS, 'Result': Result
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        MGS = data['MGS']
        Flag = data['Flag']
        ExelineS = data['Exel']
        ExecutableS = data['Exec']
        sbflsus = data['sbflsus']
        sbflmetric = data['sbflmetric']
        # sbflsus, sbflmetric = SBFL(MGS, Result, ExecutableS, ExelineS)
        # data['sbflsus'] = sbflsus
        # data['sbflmetric'] = sbflmetric
        # data['Result'] = Result
        # FAFLVariantsus = data['FAFLVariantsus_test']
        # FAFLVariantmetric = data['FAFLVariantmetric_test']
        # sus = data['sus']
        # staDe = data['staDe']
        # PSsus = data['PSsus']
        # MMsus = data['MMsus']
        # percent = data['percent']
        # FAsus, FAmetric, _, _, _, _ = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        sus, metric, staDe = Sus(MGS, ExecutableS, ExelineS)
        FAsus, FAmetric, percent, pot, pof, pal = FaSus(MGS, ExecutableS, ExelineS, Flag)
        PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLVariantsus, FAFLVariantmetric = FaflVariantSus_test(MGS, ExecutableS, ExelineS)
        FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        MMsus, MMmetric, _, _, _, _ = MmSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        # metric = data['metric']
        # FAmetric = data['FAmetric']
        # sus, metric, staDe = Sus(MGS, ExecutableS, ExelineS)
        # data['staDe'] = staDe
        # data['sus_nofs'] = sus
        # data['metric_nofs'] = metric
        # sus_nofs, metric_nofs = Sus(MGS, ExecutableS, ExelineS)
        # FAsus_nofs, FAmetric_nofs, percent_nofs = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # sus = data['sus']
        # staDe = data['staDe']
        # FAsus = data['FAsus']
        # FAFLsus = data['FAFLsus']
        # MMsus = data['MMsus']
        # percent = data['percent']
        # FAsus, FAmetric, percent, pot, pof, pal = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLVariantsus, FAFLVariantmetric = FaflVariantSus_test(MGS, ExecutableS, ExelineS)
        # FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        # MMsus, MMmetric, _, _, _, _ = MmSus(MGS, ExecutableS, ExelineS, Flag)
        # data['MMsus'] = MMsus
        # data['MMmetric'] = MMmetric
        # data['FAFLVariantsus_nofs'] = FAFLVariantsus
        # data['FAFLVariantmetric_nofs'] = FAFLVariantmetric
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # FAsus = data['FAsus']
        # FAmetric = data['FAmetric']
        # data['FAsus_nofs'] = FAsus
        # data['FAmetric_nofs'] = FAmetric
        # data['percent_nofs'] = percent
        # data['PSsus_nofs'] = PSsus
        # data['PSmetric_nofs'] = PSmetric
        # data['FAFLsus_nofs'] = FAFLsus
        # data['FAFLmetric_nofs'] = FAFLmetric
        # data['MMsus_nofs'] = MMsus
        # data['MMmetric_nofs'] = MMmetric
        # pot = data['pot']
        # pof = data['pof']
        # pal = data['pal']
        # data['sus_nofs'] = sus_nofs
        # data['metric_nofs'] = metric_nofs
        # data['FAsus_nofs'] = FAsus_nofs
        # data['FAmetric_nofs'] = FAmetric_nofs
        # data['percent_nofs'] = percent_nofs
        # data = {
        #     'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
        #     'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
        #     'Exec': ExecutableS, 'Exel': ExelineS
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # Flag = data['Flag']
        # MGS = data['MGS']
        # sus = data['sus']
        # FAsus = data['FAsus']
        # percent = data['percent']
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # data['FAsus'] = FAsus
        # data['FAmetric'] = FAmetric
        # FAsus = data['FAsus']
        # FAmetric = data['FAmetric']
        # data['FAsus_nofs'] = FAsus
        # data['FAmetric_nofs'] = FAmetric
        # data['percent_nofs'] = percent
        # data['PSsus_nofs'] = PSsus
        # data['PSmetric_nofs'] = PSmetric
        # data['FAFLsus_nofs'] = FAFLsus
        # data['FAFLmetric_nofs'] = FAFLmetric
        # data['MMsus_nofs'] = MMsus
        # data['MMmetric_nofs'] = MMmetric
        # pot = data['pot']
        # pof = data['pof']
        # pal = data['pal']
        data['sus_nofs'] = sus
        data['metric_nofs'] = metric
        data['FAsus_nofs'] = FAsus
        data['FAmetric_nofs'] = FAmetric
        data['percent_nofs'] = percent
        data['PSsus_nofs'] = PSsus
        data['PSmetric_nofs'] = PSmetric
        data['FAFLsus_nofs'] = FAFLsus
        data['FAFLmetric_nofs'] = FAFLmetric
        data['MMsus_nofs'] = MMsus
        data['MMmetric_nofs'] = MMmetric
        json_str = json.dumps(data)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'w') as f:
            json.dump(json_str, f)
        # row = eval('getMetrics_3')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, Flag, percent, pot, pof, pal, staDe)  # +sys.argv[1][-1]
        row = eval('getMetrics_3_test')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, sbflsus, Flag, percent,
                                        pot, pof, pal, staDe)  # +sys.argv[1][-1]
        # row = eval('getMetrics_4_test')(row, ws, mu, FAFLVariantsus, Flag)
    wb.save(path)
