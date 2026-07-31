from KNN import *
import json
import os
import csv
from mutants.Mutant1 import *
from mutants.Mutant2 import *
from mutants.Mutant4 import *
import coverage
from publicFun import *
from openpyxl import load_workbook
random.seed(1)


# def getOriginalInput():
#     source_case_set = []
#     with open('/Applications/work/data/MT/FAILTIM/Result/knn/iris.data', 'rt') as csvfile:
#         lines = csv.reader(csvfile)
#         dataset = list(lines)
#     for k in range(1000):
#         trainingSet = random.sample(dataset, int(len(dataset) * 0.6))
#         testSet = copy.deepcopy(dataset)
#         for i in range(len(trainingSet)):
#             testSet.remove(trainingSet[i])  # 被测样本
#         testSet = random.sample(testSet, 1)
#         source_case_set.append([trainingSet, testSet])
#
#     # 随机取100个测试用例
#     random_input = random.sample(source_case_set, 100)
#     data = {
#             'source_case_set': source_case_set,
#             'random_input': random_input
#     }
#     json_str = json.dumps(data)
#     with open('/Applications/work/data/MT/MFT/' + string + '/OriginalInput.json', 'w') as f:
#         json.dump(json_str, f)
#     return source_case_set, random_input
#
#
# def FailureRate(dynamic):
#     # 把SourceCases读出来
#     Result = []
#     with open('/Applications/work/data/MT/MFT/'+string+'/OriginalInput.json', 'r') as load_f:
#         data = json.load(load_f)
#     data = json.loads(data)
#     source_case_set = data['source_case_set']
#     for i in range(len(source_case_set)):
#         knn = KNN()
#         knn.setInput(source_case_set[i][0], source_case_set[i][1])  # oracle
#         result_s_a = knn.getPredications()
#         dynamic.setInput(source_case_set[i][0], source_case_set[i][1])  # oracle
#         result_s_m = dynamic.getPredications()
#         if result_s_a == result_s_m:
#             Result.append(0)
#         else:
#             Result.append(1)
#     FR = round(Result.count(1) / len(Result) * 100, 2)
#     return FR
#
#
# def riskIndex(argv, dynamic):
#     MGS1 = []  # 原始
#     MGS2 = []  # 新增
#     Result1 = []
#     Result2 = []
#     testcase1 = []  # 原始
#     testcase2 = []  # 新增
#     follow1 = []  # 数字代表层数的大小
#     source_case = argv.copy()
#     testcase1.append(source_case)
#     MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
#     for i in range(len(follow_case1)):
#         testcase1.append(follow_case1[i])
#     MGS1.append(MG)
#     for i in range(len(follow_case1)):  # t2t3t4
#         MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
#         follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
#         MGS1.append(MG)
#         for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
#             testcase1.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...
#
#     for i in range(len(follow1)):
#         mgs1 = []
#         ts1 = []
#         for j in range(len(follow1[i])):
#             mgs2 = []
#             ts2 = [follow1[i][j]]
#             MG, follow_case3 = MTG(follow1[i][j], dynamic)  # t5t14...
#             mgs2.append(MG)
#             follow2 = []
#             for k in range(len(follow_case3)):  # t14t15t16
#                 MG, follow_case4 = MTG(follow_case3[k], dynamic)  # t14t17...
#                 follow2.append(follow_case4)
#                 mgs2.append(MG)
#                 ts2.append(follow_case3[k])  # t14t15t16
#
#             mgs1.append(mgs2)
#
#             for n in range(len(follow2)):
#                 for m in range(len(follow2[n])):
#                     ts2.append(follow2[n][m])  # # t17t18t19...
#             ts1.append(ts2)
#
#         MGS2.append(mgs1)
#         testcase2.append(ts1)
#
#     # MG统计完, testcase统计完
#     for i in range(len(testcase1)):
#         knn = KNN()
#         knn.setInput(testcase1[i][0], testcase1[i][1])  # oracle
#         result_s_a = knn.getPredications()
#         dynamic.setInput(testcase1[i][0], testcase1[i][1])  # oracle
#         result_s_m = dynamic.getPredications()
#         if result_s_a[0] == result_s_m[0]:
#             Result1.append(0)
#         else:
#             Result1.append(1)
#
#     # 去掉巧合满足性
#     for i in range(len(MGS1)):
#         for j in range(len(MGS1[i])):
#             if MGS1[i][j] == 0 and (Result1[i] or Result1[i * len(MGS1[0]) + j + 1]):  # 如果satisfied
#                 MGS1[i][j] = 3
#
#     # MG统计完, testcase统计完
#     for i in range(len(testcase2)):
#         result1 = []
#         for j in range(len(testcase2[i])):
#             result2 = []
#             for k in range(len(testcase2[i][j])):
#                 knn = KNN()
#                 knn.setInput(testcase2[i][j][k][0], testcase2[i][j][k][1])  # oracle
#                 result_s_a = knn.getPredications()
#                 dynamic.setInput(testcase2[i][j][k][0], testcase2[i][j][k][1])  # oracle
#                 result_s_m = dynamic.getPredications()
#                 if result_s_a[0] == result_s_m[0]:
#                     result2.append(0)
#                 else:
#                     result2.append(1)
#             result1.append(result2)
#         Result2.append(result1)
#
#     # 去掉巧合满足性
#     for i in range(len(MGS2)):
#         for j in range(len(MGS2[i])):
#             for k in range(len(MGS2[i][j])):
#                 if MGS2[i][j][k] == 0 and (Result2[i][j][j] or Result2[i][j][j * len(MGS1[0]) + k]):  # 如果satisfied
#                     MGS2[i][j][k] = 3
#
#     MGS = [MGS1, MGS2]
#     Result = [Result1, Result2]
#
#     SMGS = copy.deepcopy(MGS)
#
#     # 随机去除一些MG
#     for i in range(1, len(MGS1)):  # 第一组不变
#         t = random.randint(1, len(MGS1[i])-1)  # 去几个
#         a = [n for n in range(len(MGS1[i]))]
#         random.shuffle(a)
#         b = a[:t]
#         for j in b:
#             MGS1[i][j] = 4
#
#     return MGS, Result, SMGS

def getInput(argv, dynamic):
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...
    return testcase


def statements(argv):
    testcase = argv.copy()
    filename = 'Original.py'
    Exelines = []
    for i in range(len(testcase)):
        cov = coverage.coverage()
        cov.start()
        knn = KNN()
        knn.setInput(testcase[i][0], testcase[i][1])  # oracle
        result_s_a = knn.getPredications()
        cov.stop()
        numlist = cov.analysis(filename)
        executable = numlist[1]
        exelist = list(set(numlist[1]) - set(numlist[2]))
        exelist.sort()
        Exelines.append(exelist)
    return Exelines, executable


def riskIndex(argv, dynamic):
    MGS = []  # 原始
    Result = []
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    MGS.append(MG)
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        MGS.append(MG)
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...

    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        knn = KNN()
        knn.setInput(testcase[i][0], testcase[i][1])  # oracle
        result_s_a = knn.getPredications()
        dynamic.setInput(testcase[i][0], testcase[i][1])  # oracle
        result_s_m = dynamic.getPredications()
        if result_s_a[0] == result_s_m[0]:
            Result.append(0)
        else:
            Result.append(1)


    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    return MGS, Result


if __name__ == '__main__':
    project = 'MSBF'
    row = 1
    string = 'KNN'
    path = '/Applications/work/data/MT/' + project + '/Result/result8_nofs.xlsx'  # '+sys.argv[1][:-1]+'
    wb = load_workbook(path)
    # source_case_set, random_input = getOriginalInput()
    if string not in wb.sheetnames:
        ws = wb.create_sheet(string)
    del wb[string]
    ws = wb.create_sheet(string)
    MG_set = []
    M_set = []
    Sus_set = []
    Flag_set = []
    # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/'+string+'/OriginalInput.json', 'r') as load_f:
    #     data = json.load(load_f)
    # data = json.loads(data)
    # random_input = data['random_input']
    # testcases = []
    # for i in random_input:
    #     testcases.append(getInput(i, KNN()))
    # # data = {
    # #     'testcases': testcases
    # # }
    # # json_str = json.dumps(data)
    # # with open('/Applications/work/data/MT/'+project+'/' + string + '/Input.json', 'w') as f:
    # #     json.dump(json_str, f)
    # # with open('/Applications/work/data/MT/'+project+'/' + string + '/Input.json', 'r') as f:
    # #     data = json.load(f)
    # # data = json.loads(data)
    # # testcases = data['testcases']
    # for i in testcases:
    #     Exelines, Executable = statements(i)
    #     ExelinesS.append(Exelines)
    # data = {
    #     'Exec': Executable, 'Exel': ExelinesS
    # }
    # json_str = json.dumps(data)
    # with open('/Applications/work/data/MT/'+project+'/' + string + '/statements.json', 'w') as f:
    #     json.dump(json_str, f)
    #
    for mu in [1, 2, 4]:
        # dynamic = eval("Mutant" + str(mu))()
        # MGS = []
        # Result = []
        # for i in range(100):
        #     MG, result = riskIndex(random_input[i], dynamic)
        #     MGS.append(MG)
        #     Result.append(result)
        # with open('/Applications/work/data/MT/'+project+'/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # SMGS = data['SMGS']
        # MGS = data['MGS']
        # Flag = data['Flag']
        # with open('/Users/rendaixu/OneDrive/data/MT/STVR/' + string + '/statements{}.json'.format(mu), 'r') as f:
        #     data = json.load(f)
        # data = json.loads(data)
        # ExelineS = data['Exel']
        # ExecutableS = data['Exec']
        # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # MGS = data['MGS']
        # Flag = [0] * len(ExecutableS)
        # if mu == 1:
        #     Flag[ExecutableS.index(31)] = 1
        # elif mu == 2:
        #     Flag[ExecutableS.index(30)] = 1
        # elif mu == 4:
        #     Flag[ExecutableS.index(31)] = 1
        # # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        # #           'r') as load_f:
        # #     data = json.load(load_f)
        # # data = json.loads(data)
        # # MGS = data['MGS']
        # # Flag = data['Flag']
        # # mSus = data['mSus']
        # # metric = data['metric']
        # sus, metric = Sus(MGS, ExecutableS, ExelineS)
        # FAsus, FAmetric, percent = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # data = {
        #     'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
        #     'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
        #     'Exec': ExecutableS, 'Exel': ExelineS
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        MGS = data['MGS']
        Flag = data['Flag']
        ExelineS = data['Exel']
        ExecutableS = data['Exec']
        sbflsus = data['sbflsus']
        sbflmetric = data['sbflmetric']
        # sbflsus, sbflmetric = SBFL(MGS, Result, ExecutableS, ExelineS)
        # data['sbflsus'] = sbflsus
        # data['sbflmetric'] = sbflmetric
        # data['Result'] = Result
        # FAFLVariantsus = data['FAFLVariantsus_test']
        # FAFLVariantmetric = data['FAFLVariantmetric_test']
        # sus = data['sus']
        # staDe = data['staDe']
        # PSsus = data['PSsus']
        # MMsus = data['MMsus']
        # percent = data['percent']
        # FAsus, FAmetric, _, _, _, _ = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        sus, metric, staDe = Sus(MGS, ExecutableS, ExelineS)
        FAsus, FAmetric, percent, pot, pof, pal = FaSus(MGS, ExecutableS, ExelineS, Flag)
        PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLVariantsus, FAFLVariantmetric = FaflVariantSus_test(MGS, ExecutableS, ExelineS)
        FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        MMsus, MMmetric, _, _, _, _ = MmSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        # metric = data['metric']
        # FAmetric = data['FAmetric']
        # sus, metric, staDe = Sus(MGS, ExecutableS, ExelineS)
        # data['staDe'] = staDe
        # data['sus_nofs'] = sus
        # data['metric_nofs'] = metric
        # sus_nofs, metric_nofs = Sus(MGS, ExecutableS, ExelineS)
        # FAsus_nofs, FAmetric_nofs, percent_nofs = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # sus = data['sus']
        # staDe = data['staDe']
        # FAsus = data['FAsus']
        # FAFLsus = data['FAFLsus']
        # MMsus = data['MMsus']
        # percent = data['percent']
        # FAsus, FAmetric, percent, pot, pof, pal = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLVariantsus, FAFLVariantmetric = FaflVariantSus_test(MGS, ExecutableS, ExelineS)
        # FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        # MMsus, MMmetric, _, _, _, _ = MmSus(MGS, ExecutableS, ExelineS, Flag)
        # data['MMsus'] = MMsus
        # data['MMmetric'] = MMmetric
        # data['FAFLVariantsus_nofs'] = FAFLVariantsus
        # data['FAFLVariantmetric_nofs'] = FAFLVariantmetric
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # FAsus = data['FAsus']
        # FAmetric = data['FAmetric']
        # data['FAsus_nofs'] = FAsus
        # data['FAmetric_nofs'] = FAmetric
        # data['percent_nofs'] = percent
        # data['PSsus_nofs'] = PSsus
        # data['PSmetric_nofs'] = PSmetric
        # data['FAFLsus_nofs'] = FAFLsus
        # data['FAFLmetric_nofs'] = FAFLmetric
        # data['MMsus_nofs'] = MMsus
        # data['MMmetric_nofs'] = MMmetric
        # pot = data['pot']
        # pof = data['pof']
        # pal = data['pal']
        # data['sus_nofs'] = sus_nofs
        # data['metric_nofs'] = metric_nofs
        # data['FAsus_nofs'] = FAsus_nofs
        # data['FAmetric_nofs'] = FAmetric_nofs
        # data['percent_nofs'] = percent_nofs
        # data = {
        #     'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
        #     'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
        #     'Exec': ExecutableS, 'Exel': ExelineS
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # Flag = data['Flag']
        # MGS = data['MGS']
        # sus = data['sus']
        # FAsus = data['FAsus']
        # percent = data['percent']
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # data['FAsus'] = FAsus
        # data['FAmetric'] = FAmetric
        # FAsus = data['FAsus']
        # FAmetric = data['FAmetric']
        # data['FAsus_nofs'] = FAsus
        # data['FAmetric_nofs'] = FAmetric
        # data['percent_nofs'] = percent
        # data['PSsus_nofs'] = PSsus
        # data['PSmetric_nofs'] = PSmetric
        # data['FAFLsus_nofs'] = FAFLsus
        # data['FAFLmetric_nofs'] = FAFLmetric
        # data['MMsus_nofs'] = MMsus
        # data['MMmetric_nofs'] = MMmetric
        # pot = data['pot']
        # pof = data['pof']
        # pal = data['pal']
        data['sus_nofs'] = sus
        data['metric_nofs'] = metric
        data['FAsus_nofs'] = FAsus
        data['FAmetric_nofs'] = FAmetric
        data['percent_nofs'] = percent
        data['PSsus_nofs'] = PSsus
        data['PSmetric_nofs'] = PSmetric
        data['FAFLsus_nofs'] = FAFLsus
        data['FAFLmetric_nofs'] = FAFLmetric
        data['MMsus_nofs'] = MMsus
        data['MMmetric_nofs'] = MMmetric
        json_str = json.dumps(data)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'w') as f:
            json.dump(json_str, f)
        # row = eval('getMetrics_3')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, Flag, percent, pot, pof, pal, staDe)  # +sys.argv[1][-1]
        row = eval('getMetrics_3_test')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, sbflsus, Flag, percent,
                                        pot, pof, pal, staDe)  # +sys.argv[1][-1]
        # row = eval('getMetrics_4_test')(row, ws, mu, FAFLVariantsus, Flag)
    wb.save(path)
