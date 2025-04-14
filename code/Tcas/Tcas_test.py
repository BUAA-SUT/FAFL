from coverage import CoveragePlugin
import os
from Tcas import *
import coverage
import json
from publicFun import *
from Mutants import *
from Original import *
from openpyxl import load_workbook
random.seed(1)


# def getOriginalInput():
#     source_case_set = []
#     path = r"/Applications/work/code/project/tcas/testplans.alt/universe.txt"
#     with open(path, 'r') as f:
#         list_read = f.readlines()
#     for i in range(len(list_read)):
#         test_case = []
#         list2 = list_read[i].split()
#         for j in range(len(list2)):
#             test_case.append(int(list2[j]))  # 字符串转整型
#         # 获取符合MR要求的源测试用例
#         if test_case[6] <= 3:
#             source_case_set.append(test_case)
#         if len(source_case_set) >= 1000:
#             break
#     # 随机取100个测试用例
#     random_input = random.sample(source_case_set, 100)
#     data = {
#             'source_case_set': source_case_set,
#             'random_input': random_input
#     }
#     json_str = json.dumps(data)
#     with open('/Applications/work/data/MT/MFT/' + string + '/OriginalInput.json', 'w') as f:
#         json.dump(json_str, f)
#     return source_case_set, random_input
#
#
def FailureRate(dynamic):
    # 把SourceCases读出来
    Result = []
    with open('/Applications/work/data/MT/'+project+'/'+string+'/Input.json', 'r') as load_f:
        data = json.load(load_f)
    data = json.loads(data)
    source_case_set = data['testcases']
    for i in range(len(source_case_set)):
        for j in range(len(source_case_set[i])):
            result_s_a = TCAS().Tcas(source_case_set[i][j])  # oracle
            result_s_m = dynamic.Tcas(source_case_set[i][j])
            if result_s_a == result_s_m:
                Result.append(0)
            else:
                Result.append(1)
    FR = round(Result.count(1) / len(Result) * 100, 2)
    return FR

#

def getInput(argv, dynamic):
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    MGS = []
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    MGS.append(MG)
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        MGS.append(MG)
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...
    return testcase


def statements(argv, mu):
    testcase = argv.copy()
    filename = 'Mutant{}.py'.format(mu)
    Exelines = []
    executable = []
    for i in range(len(testcase)):
        result_s_a = eval('Mutant{}()'.format(mu)).Tcas(testcase[i])
        # cov = coverage.coverage()
        # cov.start()
        # result_s_a = eval('Mutant{}()'.format(mu)).Tcas(testcase[i])  # oracle
        # cov.stop()
        # numlist = cov.analysis(filename)
        # executable = numlist[1]
        # exelist = list(set(numlist[1]) - set(numlist[2]))
        # exelist.sort()
        # Exelines.append(exelist)
    return Exelines, executable


def riskIndex(argv, dynamic):
    MGS = []  # 原始
    Result = []
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    MGS.append(MG)
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        MGS.append(MG)
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...

    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = TCAS().Tcas(testcase[i])  # oracle
        result_s_m = dynamic.Tcas(testcase[i])
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)

    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    return MGS, Result


if __name__ == '__main__':
    project = 'MSBF'
    row = 1
    string = 'Tcas'
    path = '/Applications/work/data/MT/' + project + '/Result/result8_nofs.xlsx'  # '+sys.argv[1][:-1]+'
    wb = load_workbook(path)
    # source_case_set, random_input = getOriginalInput()
    if string not in wb.sheetnames:
        ws = wb.create_sheet(string)
    del wb[string]
    ws = wb.create_sheet(string)
    FR = []
    MG_set = []
    M_set = []
    Sus_set = []
    Flag_set = []
    # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/'+string+'/OriginalInput.json', 'r') as load_f:
    #     data = json.load(load_f)
    # data = json.loads(data)
    # random_input = data['random_input']
    # testcases = []
    # for i in random_input:
    #     testcases.append(getInput(i, TCAS()))
    # data = {
    #     'testcases': testcases
    # }
    # json_str = json.dumps(data)
    # with open('/Applications/work/data/MT/'+project+'/' + string + '/Input.json', 'w') as f:
    #     json.dump(json_str, f)
    # with open('/Applications/work/data/MT/'+project+'/' + string + '/Input.json', 'r') as f:
    #     data = json.load(f)
    # data = json.loads(data)
    # testcases = data['testcases']
    # ExelinesS = []
    # for i in testcases:
    #     Exelines, Executable = statements(i)
    #     ExelinesS.append(Exelines)
    # data = {
    #     'Exec': Executable, 'Exel': ExelinesS
    # }
    # json_str = json.dumps(data)
    # with open('/Applications/work/data/MT/'+project+'/' + string + '/statements.json', 'w') as f:
    #     json.dump(json_str, f)
    #
    # with open('/Users/rendaixu/OneDrive/data/MT/STVR/' + string + '/statements.json', 'r') as f:
    #     data = json.load(f)
    # data = json.loads(data)
    # ExelinesS = data['Exel']
    # Executable = data['Exec']
    # with open('/Users/rendaixu/OneDrive/data/MT/STVR/' + string + '/Input.json', 'r') as f:
    #     data = json.load(f)
    # data = json.loads(data)
    # testcases = data['testcases']
    # for mu in range(1, 21):
    #     ExelineS = []
    #     # Executable = []
    #     for i in testcases:
    #         Exelines, Executable = statements(i, mu)
    #         ExelineS.append(Exelines)
    #     data = {
    #         'Exec': Executable, 'Exel': ExelineS
    #     }
    #     json_str = json.dumps(data)
    #     with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/statements{}.json'.format(mu), 'w') as f:
    #         json.dump(json_str, f)
    # ExelinesSet = []
    for mu in range(1, 21):
        # dynamic = TCASFactory("Mutant" + str(mu)).getTCAS()
    #     a = FailureRate(dynamic)
    #     FR.append(a)
    #     MGS = []
    #     Result = []
    #     for i in range(100):
    #         MG, result = riskIndex(random_input[i], dynamic)
    #         MGS.append(MG)
    #         Result.append(result)
    #     Executable_set.append(ExecutableS)
    #     print(mu)
    #     with open('/Users/rendaixu/OneDrive/data/MT/STVR/' + string + '/statements' + str(mu) + '.json', 'r') as load_f:
    #         data = json.load(load_f)
    #     data = json.loads(data)
    #     ExecutableS = data['Exec']
    #     ExelineS = data['Exel']
    #     # ExelinesSet.append(ExelinesS)
    #     # Flag = data['flag']
    #     # ExecutableS_set.append(ExecutableS)
    #     # Flag_set.append(Flag)
    #     Flag = [0] * len(ExecutableS)
    #     if mu == 1:
    #         Flag[ExecutableS.index(19)] = 1
    #     elif mu == 2:
    #         Flag[ExecutableS.index(16)] = 1
    #     elif mu == 3:
    #         Flag[ExecutableS.index(59)] = 1
    #     elif mu == 4:
    #         Flag[ExecutableS.index(63)] = 1
    #     elif mu == 5:
    #         Flag[ExecutableS.index(7)] = 1
    #     elif mu == 6:
    #         Flag[ExecutableS.index(61)] = 1
    #     elif mu == 7:
    #         Flag[ExecutableS.index(24)] = 1
    #     elif mu == 8:
    #         Flag[ExecutableS.index(59)] = 1
    #     elif mu == 9:
    #         Flag[ExecutableS.index(7)] = 1
    #     elif mu == 10:
    #         Flag[ExecutableS.index(60)] = 1
    #     elif mu == 11:
    #         Flag[ExecutableS.index(21)] = 1
    #     elif mu == 12:
    #         Flag[ExecutableS.index(34)] = 1
    #     elif mu == 13:
    #         Flag[ExecutableS.index(19)] = 1
    #     elif mu == 14:
    #         Flag[ExecutableS.index(29)] = 1
    #     elif mu == 15:
    #         Flag[ExecutableS.index(58)] = 1
    #     elif mu == 16:
    #         Flag[ExecutableS.index(58)] = 1
    #     elif mu == 17:
    #         Flag[ExecutableS.index(13)] = 1
    #     elif mu == 18:
    #         Flag[ExecutableS.index(16)] = 1
    #     elif mu == 19:
    #         Flag[ExecutableS.index(13)] = 1
    #     elif mu == 20:
    #         Flag[ExecutableS.index(16)] = 1
    #     with open('/Users/rendaixu/OneDrive/data/MT/STVR/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
    #         data = json.load(load_f)
    #     data = json.loads(data)
    #     MGS = data['MGS']
    #     sus, metric = Sus(MGS, ExecutableS, ExelineS)
    #     FAsus, FAmetric, percent = FaSus(MGS, ExecutableS, ExelineS, Flag)
    #     data = {
    #         'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
    #         'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
    #         'Exec': ExecutableS, 'Exel': ExelineS
    #     }
    #     json_str = json.dumps(data)
    #     with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
    #               'w') as f:
    #         json.dump(json_str, f)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        MGS = data['MGS']
        Flag = data['Flag']
        ExelineS = data['Exel']
        ExecutableS = data['Exec']
        sbflsus = data['sbflsus']
        sbflmetric = data['sbflmetric']
        # sbflsus, sbflmetric = SBFL(MGS, Result, ExecutableS, ExelineS)
        # data['sbflsus'] = sbflsus
        # data['sbflmetric'] = sbflmetric
        # data['Result'] = Result
        # FAFLVariantsus = data['FAFLVariantsus_test']
        # FAFLVariantmetric = data['FAFLVariantmetric_test']
        # sus = data['sus']
        # staDe = data['staDe']
        # PSsus = data['PSsus']
        # MMsus = data['MMsus']
        # percent = data['percent']
        # FAsus, FAmetric, _, _, _, _ = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        sus, metric, staDe = Sus(MGS, ExecutableS, ExelineS)
        FAsus, FAmetric, percent, pot, pof, pal = FaSus(MGS, ExecutableS, ExelineS, Flag)
        PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLVariantsus, FAFLVariantmetric = FaflVariantSus_test(MGS, ExecutableS, ExelineS)
        FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        MMsus, MMmetric, _, _, _, _ = MmSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        # metric = data['metric']
        # FAmetric = data['FAmetric']
        # sus, metric, staDe = Sus(MGS, ExecutableS, ExelineS)
        # data['staDe'] = staDe
        # data['sus_nofs'] = sus
        # data['metric_nofs'] = metric
        # sus_nofs, metric_nofs = Sus(MGS, ExecutableS, ExelineS)
        # FAsus_nofs, FAmetric_nofs, percent_nofs = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # sus = data['sus']
        # staDe = data['staDe']
        # FAsus = data['FAsus']
        # FAFLsus = data['FAFLsus']
        # MMsus = data['MMsus']
        # percent = data['percent']
        # FAsus, FAmetric, percent, pot, pof, pal = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLVariantsus, FAFLVariantmetric = FaflVariantSus_test(MGS, ExecutableS, ExelineS)
        # FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        # MMsus, MMmetric, _, _, _, _ = MmSus(MGS, ExecutableS, ExelineS, Flag)
        # data['MMsus'] = MMsus
        # data['MMmetric'] = MMmetric
        # data['FAFLVariantsus_nofs'] = FAFLVariantsus
        # data['FAFLVariantmetric_nofs'] = FAFLVariantmetric
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # FAsus = data['FAsus']
        # FAmetric = data['FAmetric']
        # data['FAsus_nofs'] = FAsus
        # data['FAmetric_nofs'] = FAmetric
        # data['percent_nofs'] = percent
        # data['PSsus_nofs'] = PSsus
        # data['PSmetric_nofs'] = PSmetric
        # data['FAFLsus_nofs'] = FAFLsus
        # data['FAFLmetric_nofs'] = FAFLmetric
        # data['MMsus_nofs'] = MMsus
        # data['MMmetric_nofs'] = MMmetric
        # pot = data['pot']
        # pof = data['pof']
        # pal = data['pal']
        # data['sus_nofs'] = sus_nofs
        # data['metric_nofs'] = metric_nofs
        # data['FAsus_nofs'] = FAsus_nofs
        # data['FAmetric_nofs'] = FAmetric_nofs
        # data['percent_nofs'] = percent_nofs
        # data = {
        #     'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
        #     'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
        #     'Exec': ExecutableS, 'Exel': ExelineS
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # Flag = data['Flag']
        # MGS = data['MGS']
        # sus = data['sus']
        # FAsus = data['FAsus']
        # percent = data['percent']
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # data['FAsus'] = FAsus
        # data['FAmetric'] = FAmetric
        # FAsus = data['FAsus']
        # FAmetric = data['FAmetric']
        # data['FAsus_nofs'] = FAsus
        # data['FAmetric_nofs'] = FAmetric
        # data['percent_nofs'] = percent
        # data['PSsus_nofs'] = PSsus
        # data['PSmetric_nofs'] = PSmetric
        # data['FAFLsus_nofs'] = FAFLsus
        # data['FAFLmetric_nofs'] = FAFLmetric
        # data['MMsus_nofs'] = MMsus
        # data['MMmetric_nofs'] = MMmetric
        # pot = data['pot']
        # pof = data['pof']
        # pal = data['pal']
        data['sus_nofs'] = sus
        data['metric_nofs'] = metric
        data['FAsus_nofs'] = FAsus
        data['FAmetric_nofs'] = FAmetric
        data['percent_nofs'] = percent
        data['PSsus_nofs'] = PSsus
        data['PSmetric_nofs'] = PSmetric
        data['FAFLsus_nofs'] = FAFLsus
        data['FAFLmetric_nofs'] = FAFLmetric
        data['MMsus_nofs'] = MMsus
        data['MMmetric_nofs'] = MMmetric
        json_str = json.dumps(data)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'w') as f:
            json.dump(json_str, f)
        # row = eval('getMetrics_3')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, Flag, percent, pot, pof, pal, staDe)  # +sys.argv[1][-1]
        row = eval('getMetrics_3_test')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, sbflsus, Flag, percent,
                                        pot, pof, pal, staDe)  # +sys.argv[1][-1]
        # row = eval('getMetrics_4_test')(row, ws, mu, FAFLVariantsus, Flag)
    wb.save(path)
