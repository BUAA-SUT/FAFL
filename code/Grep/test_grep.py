# 生成follow
# 获取正则表达式
from myutl.Utl import Utl
import linecache
from constant import constantNumber as constant
import os
import random
import subprocess
import csv
from MRs.MR import *
import time
import threading
from openpyxl import load_workbook
from publicFun import *
import json
import re


random.seed(1)


def random_select_MR(MR_list):
    """
    randomly select a MR from the MR list
    :param MR_list:
    :return: the name of selected MR
    """
    index = random.randint(0, len(MR_list) - 1)

    return MR_list[index]


def generate_source_test_case():
    file_path = os.path.join(os.path.abspath('.'), 'files', 'partition_scheme_testcases_1.2')
    file_path2 = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    alltestcases = []
    source_test_case = {}
    with open(file_path, 'r') as file:
        for aline in file:
            alltestcases.append(aline.strip())
    source_test_case_name = []
    with open(file_path2, 'r') as file:
        for aline in file:
            source_test_case_name.append(int(aline))
    for i in source_test_case_name:
        dictionary = {"input" + str(i): alltestcases[i-1]}
        source_test_case.update(dictionary)

    return source_test_case


def generate_follow_test_case(newdata):
    """
    根据选择的蜕变关系以及原始测试用例生成衍生测试用例
    :param MR_name: 　选择的蜕变关系的名称
    :param source_test_case: 原始测试用例
    :return: 衍生测试用例
    """
    # 读源测试用例
    inputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomInput.csv'
    # csvFile = open(inputdir, "r")
    # reader = csv.reader(csvFile)
    # # 建立空字典
    # sourcedata = {}
    # inputdata = []
    followdata = {}
    factory = MR_factory()
    # a = 0
    # b = 0
    # for item in reader:
    #     # 忽略第一行
    #     if reader.line_num == 1:
    #         inputdata.append([item[0], item[1]])
    #         continue
    #     try:
    #         a += 1
    #         sourcedata[item[0]] = item[1]
    #         inputdata.append([item[0], item[1]])
    #         # 确定测试用例的apply MR
    #         # index =
    #         MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
    #             replace('\'', '').replace('\'', '').strip().split(
    #                 ':')[1].replace('[', '').replace(']', '')
    #         MRs_list = MRs.split(', ')
    #         source_pattern = newdata.get(item[0])
    #         for i in MRs_list:
    #             b += 1
    #             MR_obj = factory.choose_MR(i)
    #             follow_test_case = MR_obj.generate_follow_test_case(source_pattern, a)
    #             followdata[item[0]+"_"+i[2:]] = follow_test_case
    #             inputdata.append([item[0]+"_"+i[2:], follow_test_case])
    #
    #             MRs = linecache.getline(constant.test_cases_2_mrs_path, b). \
    #                 replace('\'', '').replace('\'', '').strip().split(
    #                 ':')[1].replace('[', '').replace(']', '')
    #             MRs_list = MRs.split(', ')
    #             source_pattern = newdata.get(item[0] + "_" + i[2:])
    #             for j in MRs_list:
    #                 MR_obj = factory.choose_MR(j)
    #                 follow_test_case = MR_obj.generate_follow_test_case(source_pattern, b)
    #                 followdata[item[0] + "_" + i[2:] + "_" + j[2:]] = follow_test_case
    #                 inputdata.append([item[0] + "_" + i[2:] + "_" + j[2:], follow_test_case])
    #     except:
    #         print(item)
    #         print(reader.line_num)
    #         break
    # csvFile.close()
    inputdata = [["name", "value"]]
    file_path = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    with open(file_path, 'r') as file:
        for aline in file:
            # 确定测试用例的apply MR
            index1 = list(newdata.keys()).index('input' + aline.strip()) + 1
            MRs = linecache.getline(constant.test_cases_2_mrs_path, index1). \
                replace('\'', '').replace('\'', '').strip().split(
                ':')[1].replace('[', '').replace(']', '')
            MRs_list = MRs.split(', ')
            source_pattern = newdata.get('input' + aline.strip())
            inputdata.append(['input' + aline.strip(), source_pattern])
            for i in MRs_list:
                MR_obj = factory.choose_MR(i)
                follow_test_case = MR_obj.generate_follow_test_case(source_pattern, index1)
                followdata['input' + aline.strip() + "_" + i[2:]] = follow_test_case
                inputdata.append(['input' + aline.strip() + "_" + i[2:], follow_test_case])
                index2 = list(newdata.keys()).index('input' + aline.strip() + "_" + i[2:]) + 1
                source_pattern2 = newdata.get('input' + aline.strip() + "_" + i[2:])
                MRs2 = linecache.getline(constant.test_cases_2_mrs_path, index2). \
                    replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
                MRs_list2 = MRs2.split(', ')

                for j in MRs_list2:
                    try:
                        MR_obj = factory.choose_MR(j)
                        follow_test_case = MR_obj.generate_follow_test_case(source_pattern2, index2)
                        followdata['input' + aline.strip() + "_" + i[2:] + "_" + j[2:]] = follow_test_case
                        inputdata.append(['input' + aline.strip() + "_" + i[2:] + "_" + j[2:], follow_test_case])
                    except:
                        print(aline, i, j)
                        exit()

    csvFile = open(inputdir, "w")
    writer = csv.writer(csvFile)
    writer.writerows(inputdata)
    csvFile.close()
    return followdata


def extract_number_after_letter(s):
    # 正则表达式匹配：字母后面的数字，允许下划线存在或不存在
    match = re.search(r'[A-Za-z](\d+)', s)
    if match:
        return match.group(1)  # 提取匹配到的数字部分
    return None  # 如果没有匹配到则返回 None


def get_input():
    inputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/grep/RandomInput.csv'
    # inputdir = '/home/rdx/data/MT/MSBF/grep/RandomInput.csv'
    csvFile = open(inputdir, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    inputcase = {}
    select = []
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            result = extract_number_after_letter(item[0])
            if result not in select and len(select) <= 100:
                select.append(result)
            if len(select) > 100:
                break
            inputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    return inputcase


def get_output(newdata, mu):
    outputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/grep/RandomOutput{}.csv'.format(mu)
    # outputdir = '/home/rdx/data/MT/STVR/MSBF/RandomOutput{}.csv'.format(mu)
    outputdata = [["name", "value"]]
    file_path = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    with open(file_path, 'r') as file:
        a = 0
        for aline in file:
            a += 1
            # 确定测试用例的apply MR
            # aline = '3897'
            MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
                replace('\'', '').replace('\'', '').strip().split(
                ':')[1].replace('[', '').replace(']', '')
            MRs_list = MRs.split(', ')
            for i in MRs_list:
                a += 1
                source_pattern = newdata.get('input' + aline.strip())
                follow_pattern = newdata.get('input' + aline.strip() + "_" + i[2:])
                try:
                    if i != 'MR11' and i != 'MR9':
                        source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                         + "\" " + "./targetFiles/file.test"
                        follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                         + "\" " + "./targetFiles/file.test"
                    elif i == 'MR11':
                        source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                         + "\" " + "./targetFiles/MR11_" + aline.strip()
                        follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                         + "\" " + "./targetFiles/MR11_" + aline.strip()
                    else:
                        source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                         + "\" " + "./targetFiles/file.test"
                        follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                         + "\" " + "./targetFiles/file.test_MR9_follow"
                except:
                    print(aline, i, a)
                    exit()
                output_source = subprocess.run(source_command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                        shell=True).stdout.decode()
                output_source_list = output_source.split("\n")
                if 'not found' in output_source_list[0]:
                    output_source = 'command not found'
                output_follow = subprocess.run(follow_command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                        shell=True).stdout.decode()
                output_follow_list = output_follow.split("\n")
                if 'not found' in output_follow_list[0]:
                    output_follow = 'command not found'
                outputdata.append(['output_source' + aline.strip()+"_" + i[2:], output_source])
                outputdata.append(['output_follow' + aline.strip()+"_" + i[2:], output_follow])

                # 确定测试用例的apply MR
                # aline = '3897'
                MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
                    replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
                MRs_list = MRs.split(', ')
                for j in MRs_list:
                    source_pattern = newdata.get('input' + aline.strip() + "_" + i[2:])
                    follow_pattern = newdata.get('input' + aline.strip() + "_" + i[2:] + "_" + j[2:])
                    try:
                        if j != 'MR11' and j != 'MR9':
                            source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                             + "\" " + "./targetFiles/file.test"
                            follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                             + "\" " + "./targetFiles/file.test"
                        elif j == 'MR11':
                            source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                             + "\" " + "./targetFiles/MR11_" + aline.strip()
                            follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                             + "\" " + "./targetFiles/MR11_" + aline.strip()
                        else:
                            source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                             + "\" " + "./targetFiles/file.test"
                            follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                             + "\" " + "./targetFiles/file.test_MR9_follow"
                    except:
                        print(aline, i, j, a)
                        exit()
                    output_source = subprocess.run(source_command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                                   shell=True).stdout.decode()
                    output_source_list = output_source.split("\n")
                    if 'not found' in output_source_list[0]:
                        output_source = 'command not found'
                    output_follow = subprocess.run(follow_command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                                   shell=True).stdout.decode()
                    output_follow_list = output_follow.split("\n")
                    if 'not found' in output_follow_list[0]:
                        output_follow = 'command not found'
                    outputdata.append(['output_source' + aline.strip() + "_" + i[2:] + "_" + j[2:], output_source])
                    outputdata.append(['output_follow' + aline.strip() + "_" + i[2:] + "_" + j[2:], output_follow])

        csvFile = open(outputdir, "w+")
        writer = csv.writer(csvFile)
        writer.writerows(outputdata)
        csvFile.close()


def getTestcase(newdata, mu):
    inputdata = [["name", "value"]]
    outputdata = [["name", "value"]]
    inputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomInput.csv'
    outputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomOutput{}.csv'.format(mu)

    for i in range(len(newdata)):
        inputdata.append([list(newdata)[i], list(newdata.values())[i]])
        pattern = list(newdata.values())[i].strip()
        command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + pattern \
                  + "\" " + "./targetFiles/file.test"
        output = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, shell=True).stdout.decode()
        # with open(outputdir+'output'+list(newdata)[i][5:], 'w+') as file:
        #     file.write(output)
        outputdata.append(['output'+list(newdata)[i][5:], output])
    csvFile1 = open(inputdir, "w")
    writer = csv.writer(csvFile1)
    writer.writerows(inputdata)
    csvFile1.close()

    csvFile2 = open(outputdir, "w+")
    writer = csv.writer(csvFile2)
    writer.writerows(outputdata)
    csvFile2.close()


def verify_result_not_MR11(MR_name, output_source, output_follow):
    return MR_factory().verify_result_no_MR11(MR_name, output_source, output_follow)


def verify_result_MR11(input_index, output_source, output_follow):
    return MR_factory().verify_MR11_result(input_index, output_source, output_follow)


def get_correct_cases(num):
    """
    返回正确的cases
    :return:
    """
    file_path = os.path.join(os.path.abspath('.'), 'files', 'partition_scheme_testcases_1.2')
    correct_cases = []
    i = 0
    with open(file_path, 'r') as file:
        lines = file.readlines()
        lines2 = lines.copy()
        random.shuffle(lines2)
        for aline in lines2:
            aline2 = aline.strip()
            command = r"./Mutants/grep_v0/grep -E " + "\"" + aline2 \
                      + "\" " + "./targetFiles/grep1.dat"
            try:
                # 尝试执行命令
                output = subprocess.run(command, stdout=subprocess.PIPE, shell=True,
                                        stderr=subprocess.DEVNULL)
                if output.returncode != 0:
                    continue
            except subprocess.CalledProcessError as e:
                # 命令执行失败
                # output = e.output.decode()
                # print(f"Command '{command}' failed with output:\n{output}")
                pass
            else:
                # 命令执行成功
                # print(f"Command '{command}' succeeded with output:\n{output}")
                # pass
                if output.stdout.decode():
                    i += 1
                    correct_cases.append(lines.index(aline)+1)
            if i == num:
                break
    correct_cases.sort()
    with open('files/executed_correct_test_cases', 'w+') as file:
        for i in correct_cases:
            file.write(str(i) + '\n')
    return correct_cases


def verify_MRs(mu):
    Result = []
    outputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/grep/RandomOutput{}.csv'.format(mu)
    file_path = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    csvFile = open(outputdir, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    with open(file_path, 'r') as file:
        a = 0
        for aline in file:
            # aline = '75'
            a += 1
            v_r = []
            input_index = int(aline)
            MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
                replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
            MRs_list = MRs.split(', ')
            for i in MRs_list:
                # print(i)
                temp_i = i
                output_source = outputcase.get('output_source' + aline.strip() + '_' + temp_i[2:])
                output_follow = outputcase.get('output_follow' + aline.strip() + '_' + temp_i[2:])
                # ver = i == 'MR11'
                # print(ver)
                if i != 'MR11':
                    # print(9)
                    result = verify_result_not_MR11(i, output_source, output_follow)
                else:
                    # print(11)
                    result = verify_result_MR11(input_index, output_source, output_follow)
                v_r.append(result)
                if result:
                    print(aline.strip(), i)
                a += 1
                MRs2 = linecache.getline(constant.test_cases_2_mrs_path, a). \
                    replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
                MRs_list2 = MRs2.split(', ')
                input_index = int(aline)
                for j in MRs_list2:
                    output_source = outputcase.get('output_source' + aline.strip() + '_' + i[2:] + '_' + j[2:])
                    output_follow = outputcase.get('output_follow' + aline.strip() + '_' + i[2:] + '_' + j[2:])
                    if j != 'MR11':
                        result = verify_result_not_MR11(j, output_source, output_follow)
                    else:
                        result = verify_result_MR11(input_index, output_source, output_follow)
                    v_r.append(result)
                    if result:
                        print(aline.strip(), i, j)
            Result.append(v_r)
    return Result


def getMG(mu, originaldata, ExecutableS):
    MGS = []
    ExelinesS = []
    outputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/grep/RandomOutput{}.csv'.format(mu)
    file_path = os.path.join(os.path.abspath('.'), 'grep/files', 'executed_correct_test_cases')
    csvFile = open(outputdir, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    with open(file_path, 'r') as file:
        a = 0
        num = 0
        for aline in file:
            a += 1
            num += 1
            if mu > 30 and num > 100:
                break
            MG = []
            v_r1 = []
            Exelines = []
            exe1 = []
            input_index = int(aline)
            MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
                replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
            MRs_list = MRs.split(', ')
            for i in MRs_list:
                # print(i)
                temp_i = i
                output_source = outputcase.get('output_source' + aline.strip() + '_' + temp_i[2:])
                output_follow = outputcase.get('output_follow' + aline.strip() + '_' + temp_i[2:])
                if i != 'MR11':
                    result = verify_result_not_MR11(i, output_source, output_follow)
                else:
                    result = verify_result_MR11(input_index, output_source, output_follow)
                if result:
                    result = 1
                else:
                    result = 0
                if result == 0 and (int(originaldata.get('output_source' + aline.strip() + '_' + temp_i[2:])[-1]) or
                                    int(originaldata.get('output_follow' + aline.strip() + '_' + temp_i[2:])[-1])):
                    result = 3
                v_r1.append(result)
                candidate = originaldata.get('output_source' + aline.strip() + '_' + temp_i[2:])[:-1]
                candidate = [int(x) for x in candidate]
                indices1 = [i for i in range(len(candidate)) if candidate[i] == 1]
                candidate = originaldata.get('output_follow' + aline.strip() + '_' + temp_i[2:])[:-1]
                candidate = [int(x) for x in candidate]
                indices2 = [i for i in range(len(candidate)) if candidate[i] == 1]
                exe1.append([[ExecutableS[i] for i in indices1], [ExecutableS[i] for i in indices2]])
                a += 1
                MRs2 = linecache.getline(constant.test_cases_2_mrs_path, a). \
                    replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
                MRs_list2 = MRs2.split(', ')
                input_index = int(aline)
                v_r2 = []
                exe2 = []
                for j in MRs_list2:
                    output_source = outputcase.get('output_source' + aline.strip() + '_' + i[2:] + '_' + j[2:])
                    output_follow = outputcase.get('output_follow' + aline.strip() + '_' + i[2:] + '_' + j[2:])
                    if j != 'MR11':
                        result = verify_result_not_MR11(j, output_source, output_follow)
                    else:
                        result = verify_result_MR11(input_index, output_source, output_follow)
                    if result:
                        result = 1  # 违反了
                    else:
                        result = 0
                    if result == 0 and (int(originaldata.get('output_source' + aline.strip() + '_' + i[2:] + '_' + j[2:])[-1]) or
                                        int(originaldata.get('output_follow' + aline.strip() + '_' + i[2:] + '_' + j[2:])[-1])):
                        result = 3
                    v_r2.append(result)
                    candidate = originaldata.get('output_source' + aline.strip() + '_' + i[2:] + '_' + j[2:])[:-1]
                    candidate = [int(x) for x in candidate]
                    indices1 = [i for i in range(len(candidate)) if candidate[i] == 1]
                    candidate = originaldata.get('output_follow' + aline.strip() + '_' + i[2:] + '_' + j[2:])[:-1]
                    candidate = [int(x) for x in candidate]
                    indices2 = [i for i in range(len(candidate)) if candidate[i] == 1]
                    exe2.append([[ExecutableS[i] for i in indices1], [ExecutableS[i] for i in indices2]])
                MG.append(v_r2)
                Exelines.append(exe2)
            MG.insert(0, v_r1)
            Exelines.insert(0, exe1)
            MGS.append(MG)
            ExelinesS.append(Exelines)
    return MGS, ExelinesS


def getResult(mu, originaldata):
    Result = []
    outputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/grep/RandomOutput{}.csv'.format(mu)
    file_path = os.path.join(os.path.abspath('.'), 'grep/files', 'executed_correct_test_cases')
    csvFile = open(outputdir, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    with open(file_path, 'r') as file:
        a = 0
        num = 0
        for aline in file:
            a += 1
            num += 1
            if mu > 30 and num > 100:
                break
            re = []
            re1 = []
            MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
                replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
            MRs_list = MRs.split(', ')
            for i in MRs_list:
                temp_i = i
                try:
                    re1.append([int(originaldata.get('output_source' + aline.strip() + '_' + temp_i[2:])[-1]),
                                int(originaldata.get('output_follow' + aline.strip() + '_' + temp_i[2:])[-1])])
                except:
                    print(aline.strip())
                    print(temp_i)
                a += 1
                MRs2 = linecache.getline(constant.test_cases_2_mrs_path, a). \
                    replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
                MRs_list2 = MRs2.split(', ')
                re2 = []
                for j in MRs_list2:
                    re2.append([int(originaldata.get('output_source' + aline.strip() + '_' + i[2:] + '_' + j[2:])[-1]),
                                int(originaldata.get('output_follow' + aline.strip() + '_' + i[2:] + '_' + j[2:])[-1])])

                re.append(re2)
            re.insert(0, re1)
            Result.append(re)
    return Result


def dataTrans():
    """
    转换数据
    :return:
    """
    inputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomInput.csv'
    outputdir = './mapping relation/input'
    if os.path.exists(outputdir):
        os.remove(outputdir)
    csvFile = open(inputdir, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    inputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            inputcase[item[0]] = item[1]
            with open(outputdir, 'a') as file:
                file.write(item[1]+'\n')
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()


def verify(inputdata):
    path = os.path.join('/Users/rendaixu/Library/CloudStorage/OneDrive-个人/data/MT/STVR/grep/备份/一级follow/testcase_2_MRs')
    tool = Utl()
    file = tool.get_file_object(path)
    all_sets = {}
    row = 0
    for item in file:
        row += 1
        MRs = linecache.getline(path, row). \
            replace('\'', '').replace('\'', '').strip().split(
            ':')[1].replace('[', '').replace(']', '')
        MRs_list = MRs.split(', ')
        data = {
            item.split(':', 1)[0].strip(): MRs_list
        }
        all_sets.update(data)
    file_path = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    with open(file_path, 'r') as file:
        for aline in file:
            index1 = list(inputdata.keys()).index('input' + aline.strip()) + 1
            num = len(all_sets.get(str(index1)))
            a = 0
            while 1:
                try:
                    if 'input' + aline.strip() in list(inputdata.keys())[index1]:
                        a += 1
                        index1 += 1
                        if index1 >= len(list(inputdata.keys())):
                            break
                    else:
                        break
                except:
                    print('出错：'+aline)
                    exit()
                    break
            if not num == a:
                print(aline, all_sets.get(str(list(inputdata.keys()).index('input' + aline.strip()) + 1)))


def mutaterate(mu):
    outputdir1 = '/Users/rendaixu/OneDrive/data/MT/MSBF/grep/RandomOutput0.csv'
    outputdir2 = '/Users/rendaixu/OneDrive/data/MT/MSBF/grep/RandomOutput{}.csv'.format(mu)
    csvFile = open(outputdir1, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase1 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase1[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    csvFile = open(outputdir2, "r")
    reader = csv.reader(csvFile)
    outputcase2 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase2[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    num = 0
    outputcase1 = list(outputcase1.values())
    outputcase2 = list(outputcase2.values())
    index = []
    for i in range(len(outputcase1)):
        if not outputcase1[i] == outputcase2[i]:
            num += 1
            index.append(i)
    rate = num / (len(outputcase1))
    print(round(rate*100, 2))
    return index


def run_with_timeout(func, args, timeout):
    thread = threading.Thread(target=func, args=args)
    thread.start()
    thread.join(timeout)

    if thread.is_alive():
        print(f"Mutant {args[1]} 超时，跳过")
        return False
    return True


if __name__ == "__main__":
    # correct = get_correct_cases(1000)
    # source_test_case = generate_source_test_case()
    # inputdata = get_input()
    # timeout = 3600  # 设定阈值为**秒
    # for i in range(12, 51):
    #     run_with_timeout(get_output, (inputdata, i), timeout)
    # rate = []
    # Result = []
    # for i in range(0, 51):
    #     print(i)
    #     row = mutaterate(i)
    #     rate.append(row)
    #     result = verify_MRs(i)
    #     Result.append(result)

    # getTestcase(source_test_case, 0)  # 0代表original version
    # followdata = generate_follow_test_case(inputdata)
    # verify(inputdata)

    # source_results = []
    # source_result_path = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomOutput0/output13'
    # with open(source_result_path, 'r') as source_file:
    #     for aline in source_file:
    #         source_results.append(aline)
    # source_file.close()
    #
    # outputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/grep/RandomOutput0.csv'
    # csvFile = open(outputdir, "r")
    # reader = csv.reader(csvFile)
    # # 建立空字典
    # outputcase = {}
    # for item in reader:
    #     # 忽略第一行
    #     if reader.line_num == 1:
    #         continue
    #     try:
    #         outputcase[item[0]] = item[1]
    #     except:
    #         print(item)
    #         print(reader.line_num)
        # break

    # output_source = outputcase.get('output114')
    # output_source2 = output_source.split('\n')
    # dataTrans()
    project = 'MSBF'
    row = 1
    string = 'grep'
    path = '/Applications/work/data/MT/' + project + '/Result/result8_nofs.xlsx'  # '+sys.argv[1][:-1]+'
    wb = load_workbook(path)
    if string not in wb.sheetnames:
        ws = wb.create_sheet(string)
    del wb[string]
    ws = wb.create_sheet(string)
    MG_set = []
    M_set = []
    ExecutableS_set = []
    Sus_set = []
    Flag_set = []
    # ExelinesS = []
    remove_mu = [1, 2, 3, 4, 6, 7, 8, 9, 15, 17, 19, 21, 22, 24, 28, 32, 33, 37, 42, 44, 45, 47, 48]
    for mu in range(1, 57):  # range(1, 57)
        if mu in remove_mu:
            continue
        # originaldata = {}
        # if mu > 30:
        #     num = 3
        # else:
        #     num = 46
        # for i in range(1, num):
        #     datadir = '/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statementResult{}_{}.csv'.format(mu, i)
        #     csvFile = open(datadir, "r")
        #     reader = csv.reader(csvFile)
        #     # 建立空字典
        #     for item in reader:
        #         # if i != 1 and reader.line_num == 1:
        #         #     continue
        #         originaldata[item[0]] = item[1:]
        #     csvFile.close()
        # ExecutableS = originaldata.get('test cases')[:-1]
        # ExecutableS = [int(x) for x in ExecutableS]
        # Result = getResult(mu, originaldata)
        # MGS, ExelineS = getMG(mu, originaldata, ExecutableS)
        # Flag = [0] * len(ExecutableS)
        # if mu == 2:
        #     Flag[ExecutableS.index(8788)] = 1
        # elif mu == 5:
        #     Flag[ExecutableS.index(7607)] = 1
        # elif mu == 10:
        #     Flag[ExecutableS.index(8162)] = 1
        # elif mu == 11:
        #     Flag[ExecutableS.index(7912)] = 1
        # elif mu == 12:
        #     Flag[ExecutableS.index(7156)] = 1
        # elif mu == 13:
        #     Flag[ExecutableS.index(7607)] = 1
        # elif mu == 14:
        #     Flag[ExecutableS.index(8003)] = 1
        # elif mu == 15:
        #     Flag[ExecutableS.index(8162)] = 1
        # elif mu == 16:
        #     Flag[ExecutableS.index(7912)] = 1
        # elif mu == 17:
        #     Flag[ExecutableS.index(9701)] = 1
        # elif mu == 18:
        #     Flag[ExecutableS.index(7952)] = 1
        # elif mu == 20:
        #     Flag[ExecutableS.index(7893)] = 1
        # elif mu == 21:
        #     Flag[ExecutableS.index(7924)] = 1
        # elif mu == 22:
        #     Flag[ExecutableS.index(7931)] = 1
        # elif mu == 23:
        #     Flag[ExecutableS.index(7931)] = 1
        # elif mu == 24:
        #     Flag[ExecutableS.index(7935)] = 1
        # elif mu == 25:
        #     Flag[ExecutableS.index(7878)] = 1
        # elif mu == 26:
        #     Flag[ExecutableS.index(7883)] = 1
        # elif mu == 27:
        #     Flag[ExecutableS.index(7911)] = 1
        # elif mu == 28:
        #     Flag[ExecutableS.index(7939)] = 1
        # elif mu == 29:
        #     Flag[ExecutableS.index(7915)] = 1
        # elif mu == 30:
        #     Flag[ExecutableS.index(7916)] = 1
        # elif mu == 31:
        #     Flag[ExecutableS.index(7906)] = 1
        # elif mu == 32:
        #     Flag[ExecutableS.index(8472)] = 1
        # elif mu == 33:
        #     Flag[ExecutableS.index(8477)] = 1
        # elif mu == 34:
        #     Flag[ExecutableS.index(7087)] = 1
        # elif mu == 35:
        #     Flag[ExecutableS.index(7218)] = 1
        # elif mu == 36:
        #     Flag[ExecutableS.index(7207)] = 1
        # elif mu == 37:
        #     Flag[ExecutableS.index(7137)] = 1
        # elif mu == 38:
        #     Flag[ExecutableS.index(7265)] = 1
        # elif mu == 39:
        #     Flag[ExecutableS.index(7251)] = 1
        # elif mu == 40:
        #     Flag[ExecutableS.index(7197)] = 1
        # elif mu == 41:
        #     Flag[ExecutableS.index(7231)] = 1
        # elif mu == 42:
        #     Flag[ExecutableS.index(7128)] = 1
        # elif mu == 43:
        #     Flag[ExecutableS.index(7228)] = 1
        # elif mu == 44:
        #     Flag[ExecutableS.index(6996)] = 1
        # elif mu == 45:
        #     Flag[ExecutableS.index(7234)] = 1
        # elif mu == 46:
        #     Flag[ExecutableS.index(7216)] = 1
        # elif mu == 47:
        #     Flag[ExecutableS.index(7096)] = 1
        # elif mu == 48:
        #     Flag[ExecutableS.index(7145)] = 1
        # elif mu == 49:
        #     Flag[ExecutableS.index(7220)] = 1
        # elif mu == 50:
        #     Flag[ExecutableS.index(7218)] = 1
        # elif mu == 51:
        #     Flag[ExecutableS.index(7156)] = 1
        # elif mu == 52:
        #     Flag[ExecutableS.index(1729)] = 1
        # elif mu == 53:
        #     Flag[ExecutableS.index(7163)] = 1
        # elif mu == 54:
        #     Flag[ExecutableS.index(8708)] = 1
        # elif mu == 55:
        #     Flag[ExecutableS.index(8003)] = 1
        # elif mu == 56:
        #     Flag[ExecutableS.index(7142)] = 1
        # sus, metric = Sus_grep(MGS, ExecutableS, ExelinesS)
        # FAsus, FAmetric, percent = FaSus_grep(MGS, ExecutableS, ExelinesS, Flag)
        # data = {
        #     'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
        #     'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
        #     'Exec': ExecutableS, 'Exel': ExelinesS
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        MGS = data['MGS']
        Flag = data['Flag']
        ExelineS = data['Exel']
        ExecutableS = data['Exec']
        sbflsus = data['sbflsus']
        sbflmetric = data['sbflmetric']
        # sbflsus, sbflmetric = SBFL_grep(MGS, Result, ExecutableS, ExelineS)
        # data['sbflsus'] = sbflsus
        # data['sbflmetric'] = sbflmetric
        # data['Result'] = Result
        # data['MGS'] = MGS
        # data['ExelineS'] = ExelineS
        # FAFLVariantsus = data['FAFLVariantsus_test']
        # FAFLVariantmetric = data['FAFLVariantmetric_test']
        # sus = data['sus_nofs']
        # staDe = data['staDe']
        # sus = data['sus_nofs']
        # staDe = data['staDe']
        # FAsus = data['FAsus_nofs']
        # percent = data['percent_nofs']
        # sus_nofs, metric_nofs = Sus_grep(MGS, ExecutableS, ExelineS)
        sus, metric, staDe = Sus_grep(MGS, ExecutableS, ExelineS)
        # data['staDe'] = staDe
        # data['sus_nofs'] = sus
        # data['metric_nofs'] = metric
        # sus = data['sus']
        # staDe = data['staDe']
        # FAsus = data['FAsus']
        # FAFLsus = data['FAFLsus']
        # MMsus = data['MMsus']
        # percent = data['percent']
        # sus = data['sus']
        # staDe = data['staDe']
        # PSsus = data['PSsus']
        # MMsus = data['MMsus']
        # percent = data['percent']
        FAsus, FAmetric, percent, pot, pof, pal = FaSus_grep(MGS, ExecutableS, ExelineS, Flag)
        PSsus, PSmetric, _, _, _, _ = SBFLSus_grep(MGS, ExecutableS, ExelineS, Flag)
        # FAFLVariantsus, FAFLVariantmetric = FaflVariantSus_grep_test(MGS, ExecutableS, ExelineS)
        # data['FAFLVariantsus_nofs'] = FAFLVariantsus
        # data['FAFLVariantmetric_nofs'] = FAFLVariantmetric
        FAFLsus, FAFLmetric, _, _, _, _ = FaflSus_grep(MGS, ExecutableS, ExelineS, Flag)
        MMsus, MMmetric, _, _, _, _ = MmSus_grep(MGS, ExecutableS, ExelineS, Flag)
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # data['FAsus'] = FAsus
        # data['FAmetric'] = FAmetric
        # MMsus, MMmetric, _, _, _, _ = MmSus_grep(MGS, ExecutableS, ExelineS, Flag)
        # data['MMsus'] = MMsus
        # data['MMmetric'] = MMmetric
        # data['FAFLVariantsus_nofs'] = FAFLVariantsus
        # data['FAFLVariantmetric_nofs'] = FAFLVariantmetric
        # data['PSsus'] = PSsus
        # data['PSmetric'] = PSmetric
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # FAsus = data['FAsus']
        # FAmetric = data['FAmetric']
        # data['PSsus_nofs'] = PSsus
        # data['PSmetric_nofs'] = PSmetric
        # data['FAFLsus_nofs'] = FAFLsus
        # data['FAFLmetric_nofs'] = FAFLmetric
        # data['MMsus_nofs'] = MMsus
        # data['MMmetric_nofs'] = MMmetric
        # pot = data['pot']
        # pof = data['pof']
        # pal = data['pal']
        # data['sus_nofs'] = sus_nofs
        # data['metric_nofs'] = metric_nofs
        # data['FAsus_nofs'] = FAsus_nofs
        # data['FAmetric_nofs'] = FAmetric_nofs
        # data['percent_nofs'] = percent_nofs
        # data = {
        #     'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
        #     'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
        #     'Exec': ExecutableS, 'Exel': ExelineS
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # Flag = data['Flag']
        # MGS = data['MGS']
        # sus = data['sus']
        # FAsus = data['FAsus']
        # percent = data['percent']
        # row = eval('getMetrics_3')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, Flag, percent, pot, pof, pal, staDe)  # +sys.argv[1][-1]
        # pot = data['pot']
        # pof = data['pof']
        # pal = data['pal']
        # 删除 data['FAFLVariantsus']
        # if 'FAFLVariantsus' in data:
        #     del data['FAFLVariantsus']
        # if 'FAFLVariantmetric' in data:
        #     del data['FAFLVariantmetric']
        # if 'FAFLVariantsus2' in data:
        #     del data['FAFLVariantsus2']
        # if 'FAFLVariantmetric2' in data:
        #     del data['FAFLVariantmetric2']
        # if 'FAFLVariantsus2_nofs' in data:
        #     del data['FAFLVariantsus2_nofs']
        # if 'FAFLVariantmetric2_nofs' in data:
        #     del data['FAFLVariantmetric2_nofs']
        # if 'PSsus2' in data:
        #     del data['PSsus2']
        # if 'PSmetric2' in data:
        #     del data['PSmetric2']
        # if 'PSsus3' in data:
        #     del data['PSsus3']
        # if 'PSmetric3' in data:
        #     del data['PSmetric3']
        # # 重命名 data['FAFLVariantsus_test'] -> data['FAFLVariantsus']
        # if 'FAFLVariantsus_test' in data:
        #     data['FAFLVariantsus'] = data.pop('FAFLVariantsus_test')
        # if 'FAFLVariantmetric_test' in data:
        #     data['FAFLVariantmetric'] = data.pop('FAFLVariantmetric_test')
        data['sus_nofs'] = sus
        data['metric_nofs'] = metric
        data['FAsus_nofs'] = FAsus
        data['FAmetric_nofs'] = FAmetric
        data['percent_nofs'] = percent
        data['PSsus_nofs'] = PSsus
        data['PSmetric_nofs'] = PSmetric
        data['FAFLsus_nofs'] = FAFLsus
        data['FAFLmetric_nofs'] = FAFLmetric
        data['MMsus_nofs'] = MMsus
        data['MMmetric_nofs'] = MMmetric
        json_str = json.dumps(data)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'w') as f:
            json.dump(json_str, f)
        # row = eval('getMetrics_3')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, Flag, percent, pot, pof, pal, staDe)  # +sys.argv[1][-1]
        row = eval('getMetrics_3_test')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, sbflsus, Flag, percent,
                                        pot, pof, pal, staDe)  # +sys.argv[1][-1]
        # row = eval('getMetrics_4_test')(row, ws, mu, FAFLVariantsus, Flag)
        print(mu)
    wb.save(path)
