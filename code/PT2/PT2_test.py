import shutil
from typing import Dict

from PT2 import *
from openpyxl import load_workbook
from publicFun import *
import json
import sys
import os
import time
import zipfile
import csv
import subprocess
import threading
import multiprocessing
# import chardet
import codecs
import io
import dill
random.seed(1)

multiprocessing.set_start_method("fork", force=True)
multiprocessing.reduction.ForkingPickler = dill.Pickler


def copyFile(fileDir, tarDir, picknumber):
    pathDir = os.listdir(fileDir)  # 取图片的原始路径
    sample = random.sample(pathDir, picknumber)  # 随机选取picknumber数量的样本图片
    for i in range(len(sample)):
        shutil.copy(fileDir + sample[i], tarDir + "input{}.txt".format(i))
    return


def movefile(tarpath, tardir, oridir):
    if os.path.exists(tardir):
        if os.path.exists(tardir + '/' + tarpath):
            # os.remove(tarpath)
            pass
        else:
            shutil.move(oridir + tarpath, tardir)
    else:
        os.makedirs(tardir)
        shutil.move(oridir + tarpath, tardir)


def zip_file(src_dir):

    zip_name = src_dir + '.zip'

    z = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)

    for dirpath, dirnames, filenames in os.walk(src_dir):

        fpath = dirpath.replace(src_dir, '')

        fpath = fpath and fpath + os.sep or ''

        for filename in filenames:

            z.write(os.path.join(dirpath, filename), fpath+filename)

    z.close()


def un_zip(file_name):
    """unzip zip file"""
    zip_file = zipfile.ZipFile(file_name)
    tardir = '/Applications/work/data/MT/FS/PT/'
    if os.path.exists(file_name[:-4]):
        return
    for names in zip_file.namelist():
        zip_file.extract(names, tardir)
    zip_file.close()


# def getOriginalInput(num):
#     source_folder = '/Applications/work/data/MT/MFT/PT/input/'  # 源文件夹路径
#     target_file = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomInput.csv'  # 移动到新的文件夹路径
#     txt_files = [f for f in os.listdir(source_folder) if f.endswith('.txt')]
#     random.shuffle(txt_files)  # 随机选择 100 个文件
#     with open(target_file, 'w', newline='', encoding='utf-8') as f:
#         writer = csv.writer(f)
#         writer.writerow(["name", "value"])
#         n = 1
#         for file_name in txt_files:
#             source_file = os.path.join(source_folder, file_name)  # 原始文件路径
#             with open(source_file, 'r', encoding='utf-8') as f2:
#                 file_content = f2.read().strip()  # 读取文件内容并去除空格
#                 file_content = file_content.replace('\x00', '')
#                 if not file_content:
#                     continue  # 如果文件内容为空，则跳过该文件
#                 writer.writerow(['input{}'.format(n-1), file_content])  # 将文件名和内容写入 CSV 文件中
#                 n += 1
#             if n == num+1:
#                 break


def getOriginalInput(num):
    source_folder = '/Applications/work/data/MT/MFT/PT/input/'  # 源文件夹路径
    target_folder = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomInput'  # 移动到新的文件夹路径
    txt_files = [f for f in os.listdir(source_folder) if f.endswith('.txt')]
    random.shuffle(txt_files)
    n = 0
    for file_name in txt_files:
        source_file = os.path.join(source_folder, file_name)  # 原始文件路径
        target_file = os.path.join(target_folder, 'input{}.txt'.format(n))  # 目标文件路径
        with open(source_file, 'r') as f:
            file_content = f.read().strip()  # 读取文件内容并去除空格
            file_content = file_content.replace('\x00', '')
            if not file_content:
                continue  # 如果文件内容为空，则跳过该文件
        shutil.copy(source_file, target_file)  # 复制文件到目标文件夹中
        n += 1
        if n == num:
            break


def FailureRate(mu):
    # 把SourceCases读出来
    Result = []
    for i in range(1000):
        ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(0, i))
        original_output = MR().getResults(ts)
        program_ts = ts
        program_ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(mu, i))
        program_output = MR().getResults(program_ts)
        isViolate = MR().assertViolation(original_output, program_output)
        if isViolate:
            Result.append(1)
        else:
            Result.append(0)
    FR = round(Result.count(1) / len(Result) * 100, 2)
    return FR


def getMG(mu, mr_list, num_of_samples, originaloutput, ts, removecase):
    outputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT2/RandomOutput{}.csv'.format(mu)
    csvFile = open(outputdir, "r")
    reader = csv.reader(csvFile)
    mutateoutput = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            mutateoutput[item[0]] = item[1].split("\n")
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    MGS = []
    for i in range(num_of_samples):
        if i in removecase:
            continue
        MGs = []
        MG = [0] * len(mr_list)
        ts.setInputOutput("input{}".format(i), "output{}".format(i))
        original_output = MR().getResults(ts, mutateoutput)
        followup_ts = copy.copy(ts)
        for j in range(len(mr_list)):
            followup_ts.setInputOutput("input{}_{}".format(i, j), "output{}_{}".format(i, j))
            mr = mr_list[j]
            followup_output = mr.getResults(followup_ts, mutateoutput)
            expected_output = mr.getExpectedOutput(ts, original_output)
            isViolate = mr.assertViolation(expected_output, followup_output)
            if isViolate:
                MG[j] = 1
            elif (mutateoutput["output{}".format(i)] != originaloutput["output{}".format(i)] or
                  mutateoutput["output{}_{}".format(i, j)] != originaloutput["output{}_{}".format(i, j)]):
                MG[j] = 3
        MGs.append(MG)
        for m in range(len(mr_list)):
            MG = [0] * len(mr_list)
            ts.setInputOutput("input{}_{}".format(i, m), "output{}_{}".format(i, m))
            original_output = MR().getResults(ts, mutateoutput)
            followup_ts = copy.copy(ts)
            for n in range(len(mr_list)):
                followup_ts.setInputOutput("input{}_{}_{}".format(i, m, n), "output{}_{}_{}".format(i, m, n))
                mr = mr_list[n]
                followup_output = mr.getResults(followup_ts, mutateoutput)
                expected_output = mr.getExpectedOutput(ts, original_output)
                isViolate = mr.assertViolation(expected_output, followup_output)
                if isViolate:
                    MG[n] = 1
                elif (mutateoutput["output{}_{}".format(i, m)] != originaloutput["output{}_{}".format(i, m)] or
                      mutateoutput["output{}_{}_{}".format(i, m, n)] != originaloutput["output{}_{}_{}".format(i, m, n)]):
                    MG[n] = 3
            MGs.append(MG)
        MGS.append(MGs)
    return MGS


def get_input():
    inputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT2/RandomInput.csv'
    # inputdir = '/home/rdx/data/MT/STVR/PT2/RandomInput.csv'
    # 建立空字典
    csvFile = open(inputdir, "r")
    # reader = csv.reader(csvFile)
    reader = csv.reader((line.replace('\0', '') for line in csvFile))
    # 建立空字典
    inputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            inputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()
    return inputcase


def dataTrans(mr, num_of_samples):
    inputs_write = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomInput.csv'
    # outputs_write = '/Users/rendaixu/OneDrive/data/MT/FS/PT/RandomOutput'
    inputs_read = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomInput'
    # outputs_read = '/Applications/work/data/MT/FS/PT/RandomOutput'
    fileHeader = ["name", "value"]
    inputdata = [fileHeader]

    for i in range(num_of_samples):
        f = open(inputs_read+'/input'+str(i)+'.txt')
        data = ['input{}'.format(i), f.read()]
        inputdata.append(data)
        for j in range(mr):
            f = open(inputs_read + '/input{}_{}.txt'.format(i, j))
            data = ['input{}_{}'.format(i, j), f.read()]
            inputdata.append(data)
            for k in range(mr):
                f = open(inputs_read + '/input{}_{}_{}.txt'.format(i, j, k))
                data = ['input{}_{}_{}'.format(i, j, k), f.read()]
                inputdata.append(data)
    # 将数据存下来
    with open(inputs_write, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(inputdata)


def verify(mr, num_of_samples):
    inputs_read = '/Applications/work/data/MT/MFT/PT/RandomOutput'
    fileHeader = ["name", "value"]
    inputdata = [fileHeader]
    for i in range(num_of_samples):
        f = open(inputs_read+'/output0_'+str(i)+'.txt')
        data = ['output0_{}'.format(i), f.read()]
        inputdata.append(data)
        for j in range(mr):
            f = open(inputs_read + '/output0_{}_{}.txt'.format(i, j))
            data = ['output0_{}_{}'.format(i, j), f.read()]
            inputdata.append(data)
            for k in range(mr):
                f = open(inputs_read + '/output0_{}_{}_{}.txt'.format(i, j, k))
                data = ['output0_{}_{}_{}'.format(i, j, k), f.read()]
                inputdata.append(data)
    inputcase = get_input()
    lst1 = list(inputcase.items())
    lst2 = list(inputcase.items())
    for i in range(len(lst1)):
        if not lst1[i][1] == lst2[i][1]:
            print(i)


def getTestcase(mr_list, test_case, num_of_samples):
    # for i in range(num_of_samples):
    #     test_case.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
    #     test_case.generateRandomTestcase()
    # for i in range(num_of_samples):
    #     for j in range(len(mr_list)):
    #         mr = mr_list[j]
    #         mr.setTestCase(test_case)
    #         mr.original_ts.setInputOutput("input{}.fa".format(i), "reference{}.fa".format(i), "e{}.fa".format(i), "output{}.txt".format(i))
    #         mr.getFollowInput(j)
    for i in range(num_of_samples):
        # test_case.setInputOutput("input{}.txt".format(i), "output{}.txt".format(i))
        # test_case.generateInput()  # 生成原始测试用例
        for j in range(len(mr_list)):
            mr = mr_list[j]
            mr.setTestCase(test_case)
            mr.original_ts.setInputOutput("input{}.txt".format(i), "output{}.txt".format(i))
            mr.getFollowInput(j)
            for k in range(len(mr_list)):
                mr = mr_list[k]
                mr.setTestCase(test_case)
                mr.original_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}.txt".format(i, j))
                mr.getFollowInput(k)


# def getOutput(mu, inputcase):
#     fileHeader = ["name", "value"]
#     outputdata = [fileHeader]
#     command = './Mutants/printtokens2_v{}/print_tokens2'.format(mu)
#     lst = list(inputcase.items())
#     for i in range(len(lst)):
#         # 将输入字符串写入临时文件
#         try:
#             output = subprocess.run(command, input=lst[i][1].encode(), stdout=subprocess.PIPE,
#                                     stderr=subprocess.STDOUT, shell=True).stdout.decode()
#         except UnicodeDecodeError:
#             print("Mutant{} 编码格式问题，跳过".format(mu))
#             return 0
#         data = ['output'+lst[i][0][5:], output]
#         outputdata.append(data)
#     outputs_write = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT2/RandomOutput{}.csv'.format(mu)
#     # outputs_write = '/home/rdx/data/MT/STVR/PT2/RandomOutput{}.csv'.format(mu)
#     with open(outputs_write, 'w', newline='') as csvfile:
#         writer = csv.writer(csvfile)
#         writer.writerows(outputdata)
#     print('第{}个已完成'.format(mu))


def mutaterate(mu):
    outputdir1 = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT2/RandomOutput0.csv'
    outputdir2 = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT2/RandomOutput{}.csv'.format(mu)
    if not os.path.exists(outputdir2):
        return 0
    csvFile = open(outputdir1, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase1 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase1[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    csvFile = open(outputdir2, "r")
    reader = csv.reader(csvFile)
    outputcase2 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase2[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    num = 0
    outputcase1 = list(outputcase1.values())
    outputcase2 = list(outputcase2.values())
    index = []
    for i in range(len(outputcase1)):
        if not outputcase1[i] == outputcase2[i]:
            num += 1
            index.append(i)
    rate = num / (len(outputcase1))
    print(mu, round(rate*100, 2))
    return index


# def run_getOutput(i, inputcase, timeout):
#     p = multiprocessing.Process(target=getOutput, args=(i, inputcase))
#     p.start()
#     p.join(timeout)
#     if p.is_alive():
#         p.terminate()
#         print("Mutant{} 超时，跳过".format(i))

def target_function(processor):
    processor.getOutput()


class MutantProcessor:
    def __init__(self, mu, inputcase):
        self.mu = mu
        self.inputcase = inputcase

    def getOutput(self):
        fileHeader = ["name", "value"]
        outputdata = [fileHeader]
        command = './Mutants/printtokens2_v{}/print_tokens2'.format(self.mu)
        lst = list(self.inputcase.items())

        for i in range(len(lst)):
            # 将输入字符串写入临时文件
            try:
                output = subprocess.run(
                    command,
                    input=lst[i][1].encode(),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    shell=True
                ).stdout.decode()
            except UnicodeDecodeError:
                print("Mutant{} 编码格式问题，跳过".format(self.mu))
                return 0

            data = ['output' + lst[i][0][5:], output]
            outputdata.append(data)

        outputs_write = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT2/RandomOutput{}.csv'.format(self.mu)

        with open(outputs_write, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(outputdata)

        print('第{}个已完成'.format(self.mu))


def run_getOutput(i, inputcase, timeout):
    # 创建 MutantProcessor 类的实例
    processor = MutantProcessor(mu=i, inputcase=inputcase)

    # 创建 multiprocessing.Process 实例
    p = multiprocessing.Process(target=target_function, args=(processor,))
    p.start()

    # 等待进程完成，设置超时
    p.join(timeout)

    # 如果进程超时，终止进程
    if p.is_alive():
        p.terminate()
        p.join()  # 确保进程结束
        print("Mutant{} 超时，跳过".format(i))
    else:
        print("Mutant{} 完成".format(i))


if __name__ == "__main__":
    num_case = 100
    removecase = [3, 41, 68]
    num_mr = 11
    num_mu = 21
    # getOriginalInput(num_case)
    # dataTrans(num_mr, num_case)
    # inputcase = get_input()
    # timeout = 600  # 设定阈值为**秒
    # for i in range(50, 62):
    #     run_getOutput(i, inputcase, timeout)
    # for mu in range(50, 62):
    #     mutaterate(mu)
    # outputdir = '/Users/rendaixu/OneDrive/data/MT/MSBF/PT2/RandomOutput0.csv'
    # csvFile = open(outputdir, "r")
    # reader = csv.reader(csvFile)
    # # 建立空字典
    # originaloutput = {}
    # for item in reader:
    #     # 忽略第一行
    #     if reader.line_num == 1:
    #         continue
    #     try:
    #         originaloutput[item[0]] = item[1].split("\n")
    #     except:
    #         print(item)
    #         print(reader.line_num)
    #         break
    # csvFile.close()
    # mr_list = [MR1(), MR2(), MR3(), MR4(), MR5(), MR6(), MR7(), MR8(), MR9(), MR10(), MR11()]
    # ts = TestCase()
    # getTestcase(mr_list, ts, num_case)


    # MGS_set = []
    # SMGS_set = []
    # for mu in range(50, 62):
    #     MGS, SMGS = getMG(mu, mr_list, num_case, originaloutput, ts)
    #     MGS_set.append(MGS)
    #     SMGS_set.append(SMGS)
    #     a = 0
    #     for i in range(len(SMGS)):
    #         for j in range(1):
    #             if 1 in SMGS[i][j]:
    #                 a = 1
    #     if a == 0:
    #         print(mu)

    # verify(mr, num_of_samples)
    # myenv = MyEnv()
    # myenv.CreateWorkingDirs()
    project = 'MSBF'
    row = 1
    string = 'PT2'
    path = '/Applications/work/data/MT/' + project + '/Result/result8_nofs.xlsx'  # '+sys.argv[1][:-1]+'
    # datadir = '/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statementResult.csv'
    wb = load_workbook(path)
    if string not in wb.sheetnames:
        ws = wb.create_sheet(string)
    del wb[string]
    ws = wb.create_sheet(string)
    MG_set = []
    M_set = []
    ExecutableS_set = []
    Sus_set = []
    Flag_set = []
    # remove_mu = [2, 4, 5, 6, 10, 11]
    for mu in range(1, num_mu+1):  #num_mu+1
        # if mu not in remove_mu:
        #     continue
        # datadir = '/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statementResult{}.csv'.format(mu)
        # csvFile = open(datadir, "r")
        # reader = csv.reader(csvFile)
        # # # 建立空字典
        # originaldata = {}
        # result = {}
        # for item in reader:
        #     originaldata[item[0]] = item[1:-1]
        #     if reader.line_num == 1:
        #         continue
        #     result[item[0]] = item[-1:]
        # csvFile.close()
        # ExecutableS = originaldata.get('inputs')
        # ExecutableS = [int(x) for x in ExecutableS]
        # # ExelineS = []
        # Result = []
        # for i in range(num_case):
        #     if i in removecase:
        #         continue
        #     Exelines = []
        #     r = []
        #     candidate = originaldata.get('input{}'.format(i))
        #     candidate = [int(x) for x in candidate]
        #     indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #     Exelines.append([ExecutableS[i] for i in indices])
        #     a = result.get('input{}'.format(i))
        #     a = [int(x) for x in a]
        #     r.append(a)
        #     for j in range(num_mr):
        #         candidate = originaldata.get('input{}_{}'.format(i, j))
        #         candidate = [int(x) for x in candidate]
        #         indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #         Exelines.append([ExecutableS[i] for i in indices])
        #         a = result.get('input{}_{}'.format(i, j))
        #         a = [int(x) for x in a]
        #         r.append(a)
        #     for m in range(num_mr):
        #         for n in range(num_mr):
        #             candidate = originaldata.get('input{}_{}_{}'.format(i, m, n))
        #             candidate = [int(x) for x in candidate]
        #             indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #             Exelines.append([ExecutableS[i] for i in indices])
        #             a = result.get('input{}_{}_{}'.format(i, m, n))
        #             a = [int(x) for x in a]
        #             r.append(a)
        #     Result.append(r)
        #     ExelineS.append(Exelines)
        # # MGS = getMG(mu, mr_list, num_case, originaloutput, ts, removecase)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # MGS = data['MGS']
        # # ExecutableS = data['Exec']
        # # ExelinesS = data['Exel']
        # # Flag = data['Flag']
        # # Sus = MSlice(SMGS, ExecutableS, ExelinesS)
        # # FaSus, percent = FAILTIMSlice(SMGS, ExecutableS, ExelinesS, Flag)
        # # mSus = data['Sus']
        # # mSus2, AllF = MSlice2(SMGS, ExecutableS, ExelinesS, Flag)
        # # mSus = data['Sus']
        # Flag = [0] * len(ExecutableS)
        # if mu == 1:
        #     Flag[49] = 1
        # elif mu == 2:
        #     Flag[50] = 1
        # elif mu == 3:
        #     Flag[51] = 1
        # elif mu == 4:
        #     Flag[ExecutableS.index(185)] = 1
        # elif mu == 5:
        #     Flag[ExecutableS.index(190)] = 1
        # elif mu == 6:
        #     Flag[ExecutableS.index(190)] = 1
        # elif mu == 7:
        #     Flag[80] = 1
        # elif mu == 8:
        #     Flag[84] = 1
        # elif mu == 9:
        #     Flag[85] = 1
        # elif mu == 10:
        #     Flag[ExecutableS.index(226)] = 1
        # elif mu == 11:
        #     Flag[ExecutableS.index(226)] = 1
        # elif mu == 12:
        #     Flag[145] = 1
        # elif mu == 13:
        #     Flag[161] = 1
        # elif mu == 14:
        #     Flag[40] = 1
        # elif mu == 15:
        #     Flag[51] = 1
        # elif mu == 16:
        #     Flag[51] = 1
        # elif mu == 17:
        #     Flag[80] = 1
        # elif mu == 18:
        #     Flag[81] = 1
        # elif mu == 19:
        #     Flag[84] = 1
        # elif mu == 20:
        #     Flag[84] = 1
        # elif mu == 21:
        #     Flag[145] = 1
        # # mSus, metric = MSlice(MGS, ExecutableS, ExelinesS)
        # # mSus2, AllF, metric2 = MSlice2(MGS, ExecutableS, ExelinesS, Flag)
        # # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        # #           'r') as load_f:
        # #     data = json.load(load_f)
        # # data = json.loads(data)
        # # MGS = data['MGS']
        # # Flag = data['Flag']
        # # ExecutableS = data['Exec']
        # # ExelineS = data['Exel']
        # sus, metric = Sus(MGS, ExecutableS, ExelineS)
        # FAsus, FAmetric, percent = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # data = {
        #     'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
        #     'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
        #     'Exec': ExecutableS, 'Exel': ExelineS, 'Result': Result
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        MGS = data['MGS']
        Flag = data['Flag']
        ExelineS = data['Exel']
        ExecutableS = data['Exec']
        sbflsus = data['sbflsus']
        sbflmetric = data['sbflmetric']
        # sbflsus, sbflmetric = SBFL(MGS, Result, ExecutableS, ExelineS)
        # data['sbflsus'] = sbflsus
        # data['sbflmetric'] = sbflmetric
        # data['Result'] = Result
        # FAFLVariantsus = data['FAFLVariantsus_test']
        # FAFLVariantmetric = data['FAFLVariantmetric_test']
        # sus = data['sus']
        # staDe = data['staDe']
        # PSsus = data['PSsus']
        # MMsus = data['MMsus']
        # percent = data['percent']
        # FAsus, FAmetric, _, _, _, _ = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        sus, metric, staDe = Sus(MGS, ExecutableS, ExelineS)
        FAsus, FAmetric, percent, pot, pof, pal = FaSus(MGS, ExecutableS, ExelineS, Flag)
        PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLVariantsus, FAFLVariantmetric = FaflVariantSus_test(MGS, ExecutableS, ExelineS)
        FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        MMsus, MMmetric, _, _, _, _ = MmSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        # metric = data['metric']
        # FAmetric = data['FAmetric']
        # sus, metric, staDe = Sus(MGS, ExecutableS, ExelineS)
        # data['staDe'] = staDe
        # data['sus_nofs'] = sus
        # data['metric_nofs'] = metric
        # sus_nofs, metric_nofs = Sus(MGS, ExecutableS, ExelineS)
        # FAsus_nofs, FAmetric_nofs, percent_nofs = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # sus = data['sus']
        # staDe = data['staDe']
        # FAsus = data['FAsus']
        # FAFLsus = data['FAFLsus']
        # MMsus = data['MMsus']
        # percent = data['percent']
        # FAsus, FAmetric, percent, pot, pof, pal = FaSus(MGS, ExecutableS, ExelineS, Flag)
        # PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLVariantsus, FAFLVariantmetric = FaflVariantSus_test(MGS, ExecutableS, ExelineS)
        # FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        # MMsus, MMmetric, _, _, _, _ = MmSus(MGS, ExecutableS, ExelineS, Flag)
        # data['MMsus'] = MMsus
        # data['MMmetric'] = MMmetric
        # data['FAFLVariantsus_nofs'] = FAFLVariantsus
        # data['FAFLVariantmetric_nofs'] = FAFLVariantmetric
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # FAsus = data['FAsus']
        # FAmetric = data['FAmetric']
        # data['FAsus_nofs'] = FAsus
        # data['FAmetric_nofs'] = FAmetric
        # data['percent_nofs'] = percent
        # data['PSsus_nofs'] = PSsus
        # data['PSmetric_nofs'] = PSmetric
        # data['FAFLsus_nofs'] = FAFLsus
        # data['FAFLmetric_nofs'] = FAFLmetric
        # data['MMsus_nofs'] = MMsus
        # data['MMmetric_nofs'] = MMmetric
        # pot = data['pot']
        # pof = data['pof']
        # pal = data['pal']
        # data['sus_nofs'] = sus_nofs
        # data['metric_nofs'] = metric_nofs
        # data['FAsus_nofs'] = FAsus_nofs
        # data['FAmetric_nofs'] = FAmetric_nofs
        # data['percent_nofs'] = percent_nofs
        # data = {
        #     'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
        #     'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
        #     'Exec': ExecutableS, 'Exel': ExelineS
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # Flag = data['Flag']
        # MGS = data['MGS']
        # sus = data['sus']
        # FAsus = data['FAsus']
        # percent = data['percent']
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # data['FAsus'] = FAsus
        # data['FAmetric'] = FAmetric
        # FAsus = data['FAsus']
        # FAmetric = data['FAmetric']
        # data['FAsus_nofs'] = FAsus
        # data['FAmetric_nofs'] = FAmetric
        # data['percent_nofs'] = percent
        # data['PSsus_nofs'] = PSsus
        # data['PSmetric_nofs'] = PSmetric
        # data['FAFLsus_nofs'] = FAFLsus
        # data['FAFLmetric_nofs'] = FAFLmetric
        # data['MMsus_nofs'] = MMsus
        # data['MMmetric_nofs'] = MMmetric
        # pot = data['pot']
        # pof = data['pof']
        # pal = data['pal']
        data['sus_nofs'] = sus
        data['metric_nofs'] = metric
        data['FAsus_nofs'] = FAsus
        data['FAmetric_nofs'] = FAmetric
        data['percent_nofs'] = percent
        data['PSsus_nofs'] = PSsus
        data['PSmetric_nofs'] = PSmetric
        data['FAFLsus_nofs'] = FAFLsus
        data['FAFLmetric_nofs'] = FAFLmetric
        data['MMsus_nofs'] = MMsus
        data['MMmetric_nofs'] = MMmetric
        json_str = json.dumps(data)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'w') as f:
            json.dump(json_str, f)
        # row = eval('getMetrics_3')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, Flag, percent, pot, pof, pal, staDe)  # +sys.argv[1][-1]
        row = eval('getMetrics_3_test')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, sbflsus, Flag, percent,
                                        pot, pof, pal, staDe)  # +sys.argv[1][-1]
        # row = eval('getMetrics_4_test')(row, ws, mu, FAFLVariantsus, Flag)
    wb.save(path)
