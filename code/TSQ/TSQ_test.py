import os
# import sys
# package_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# sys.path.append(package_path)
from TSQ import *
import coverage
# import numpy as np
# print(coverage.__file__)
import json
from publicFun import *
from Original import *
from Mutant1 import *
from Mutant2 import *
from Mutant3 import *
from Mutant4 import *
from Mutant5 import *
from Mutant6 import *
from openpyxl import load_workbook
# import xlwt
random.seed(1)


# def getOriginalInput():
#     source_case_set = []
#     while 1:
#         # a = random.randint(1, 10)
#         # b = random.randint(1, 10)
#         # c = random.randint(1, 10)
#         a = round(random.uniform(1.0, 10.0), 2)
#         b = round(random.uniform(1.0, 10.0), 2)
#         c = round(random.uniform(1.0, 10.0), 2)
#         if a >= b + c or b >= a + c or c >= a + b:
#             continue
#         else:
#             triangle = [a, b, c]
#         source_case_set.append(triangle)
#         if len(source_case_set) >= 1000:
#             break
#     # 随机取100个测试用例
#     random_input = random.sample(source_case_set, 100)
#     data = {
#             'source_case_set': source_case_set,
#             'random_input': random_input
#     }
#     json_str = json.dumps(data)
#     with open('/Applications/work/data/MT/FS/'+string2+'/OriginalInputNew.json', 'w') as f:
#         json.dump(json_str, f)
#
#     return source_case_set, random_input
#
#
# def FailureRate(dynamic):
#     # 把SourceCases读出来
#     Result = []
#     with open('/Applications/work/data/MT/MFT/'+string+'/OriginalInputNew.json', 'r') as load_f:
#         data = json.load(load_f)
#     data = json.loads(data)
#     source_case_set = data['source_case_set']
#     for i in range(len(source_case_set)):
#         result_s_a = Trisquare().trisquare(source_case_set[i])  # oracle
#         result_s_m = dynamic.trisquare(source_case_set[i])
#         if result_s_a[0] == result_s_m[0]:
#             Result.append(0)
#         else:
#             Result.append(1)
#     FR = round(Result.count(1) / len(Result) * 100, 2)
#     return FR
#
#
# def getResult():
#     Result = []
#     rate = []
#     # while 1:
#     #     a = random.randint(1, 10)
#     #     b = random.randint(1, 10)
#     #     c = random.randint(1, 10)
#     #     if a >= b + c or b >= a + c or c >= a + b:
#     #         continue
#     #     else:
#     #         triangle = [a, b, c]
#     #
#     #         result = Trisquare().trisquare2(triangle)
#     #         Result.append(result)
#     #         if len(Result) >= 1000:
#     #             break
#     with open('/Applications/work/data/MT/MFT/'+string+'/OriginalInputNew.json', 'r') as load_f:
#         data = json.load(load_f)
#     data = json.loads(data)
#     source_case_set = data['source_case_set']
#     for i in range(len(source_case_set)):
#         result = Trisquare().trisquare2(source_case_set[i])  # oracle
#         Result.append(result)
#     rate1 = Result.count(1) / len(Result) * 100
#     rate2 = Result.count(2) / len(Result) * 100
#     rate3 = Result.count(3) / len(Result) * 100
#     rate4 = Result.count(4) / len(Result) * 100
#     rate5 = Result.count(5) / len(Result) * 100
#     rate6 = Result.count(6) / len(Result) * 100
#     rate.append([rate1, rate2, rate3, rate4, rate5, rate6])
#     return rate


def getInput(argv, dynamic):
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...
    return testcase


def statements(argv):
    testcase = argv.copy()
    filename = 'Original.py'
    Exelines = []
    for i in range(len(testcase)):
        cov = coverage.coverage()
        cov.start()
        result_s_a = Trisquare().trisquare(testcase[i])  # oracle
        cov.stop()
        numlist = cov.analysis(filename)
        executable = numlist[1]
        exelist = list(set(numlist[1]) - set(numlist[2]))
        exelist.sort()
        Exelines.append(exelist)
    return Exelines, executable


def riskIndex(argv, dynamic):
    MGS = []  # 原始
    Result = []
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    MGS.append(MG)
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        MGS.append(MG)
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...

    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = Trisquare().trisquare(testcase[i])  # oracle
        result_s_m = dynamic.trisquare(testcase[i])
        if result_s_a[0] == result_s_m[0]:
            Result.append(0)
        else:
            Result.append(1)

    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    return MGS, Result


# def judgeMG(MG):
#     # b = 0
#     # for i in range(len(MG)):
#     #     # a = random.sample(MG[i][0], d)
#     #     a = MG[i][0].count(1)
#     #     if a > 2:
#     #         a = 2
#     #     b += a
#     #     if b >= num:
#     #         return i  # 目前的数据满足要求
#     #     else:
#     #         pass
#
#     a = MG[0].count(1)
#     if a > 2:
#         a = 2
#     return a


if __name__ == '__main__':
    project = 'MSBF'
    row = 1
    string = 'TSQ'
    path = '/Applications/work/data/MT/'+project+'/Result/result8_nofs.xlsx'  # '+sys.argv[1][:-1]+'
    wb = load_workbook(path)
    # source_case_set, random_input = getOriginalInput()
    if string not in wb.sheetnames:
        ws = wb.create_sheet(string)
    del wb[string]
    ws = wb.create_sheet(string)
    MG_set = []
    Sus_set = []
    Flag_set = []
    # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/'+string+'/OriginalInputNew.json', 'r') as load_f:
    #     data = json.load(load_f)
    # data = json.loads(data)
    # random_input = data['random_input']
    # testcases = []
    # for i in random_input:
    #     testcases.append(getInput(i, Trisquare()))
    # data = {
    #     'testcases': testcases
    # }
    # json_str = json.dumps(data)
    # with open('/Applications/work/data/MT/'+project+'/' + string + '/Input.json', 'w') as f:
    #     json.dump(json_str, f)
    # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/Input.json', 'r') as f:
    #     data = json.load(f)
    # data = json.loads(data)
    # testcases = data['testcases']
    for mu in range(1, 6):  # 7
        # dynamic = eval("Mutant" + str(mu))()
        # MGS = []
        # Result = []
        # for i in range(100):
        #     MG, result = riskIndex(random_input[i], dynamic)
        #     MGS.append(MG)
        #     Result.append(result)

        with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        MGS = data['MGS']
        # Result = data['Result']
        Flag = data['Flag']
        ExelineS = data['Exel']
        ExecutableS = data['Exec']
        sbflsus = data['sbflsus']
        sbflmetric = data['sbflmetric']
        # FAFLVariantsus = data['FAFLVariantsus_test']
        # FAFLVariantmetric = data['FAFLVariantmetric_test']
        # sus = data['sus_nofs']
        # FAsus = data['FAsus_nofs']
        # percent = data['percent_nofs']
        # metric = data['metric']
        # FAmetric = data['FAmetric']
        # Flag = [0] * len(ExecutableS)
        # if mu == 1:
        #     Flag[ExecutableS.index(21)] = 1
        # elif mu == 2:
        #     Flag[ExecutableS.index(26)] = 1
        # elif mu == 3:
        #     Flag[ExecutableS.index(29)] = 1
        # elif mu == 4:
        #     Flag[ExecutableS.index(34)] = 1
        # elif mu == 5:
        #     Flag[ExecutableS.index(40)] = 1
        # elif mu == 6:
        #     Flag[ExecutableS.index(42)] = 1
        # sus, metric, staDe = Sus(MGS, ExecutableS, ExelineS)
        # sbflsus, sbflmetric = SBFL(MGS, Result, ExecutableS, ExelineS)
        # data['sbflsus'] = sbflsus
        # data['sbflmetric'] = sbflmetric
        # data['Result'] = Result
        # data['staDe'] = staDe
        # data['sus_nofs'] = sus
        # data['metric_nofs'] = metric
        sus, metric, staDe = Sus(MGS, ExecutableS, ExelineS)
        FAsus, FAmetric, percent, pot, pof, pal = FaSus(MGS, ExecutableS, ExelineS, Flag)
        PSsus, PSmetric, _, _, _, _ = SBFLSus(MGS, ExecutableS, ExelineS, Flag)
        # FAFLVariantsus, FAFLVariantmetric = FaflVariantSus_test(MGS, ExecutableS, ExelineS)
        FAFLsus, FAFLmetric, _, _, _, _ = FaflSus(MGS, ExecutableS, ExelineS, Flag)
        MMsus, MMmetric, _, _, _, _ = MmSus(MGS, ExecutableS, ExelineS, Flag)
        # data['MMsus'] = MMsus
        # data['MMmetric'] = MMmetric
        # data['FAFLVariantsus_nofs'] = FAFLVariantsus
        # data['FAFLVariantmetric_nofs'] = FAFLVariantmetric
        # data['FAFLsus'] = FAFLsus
        # data['FAFLmetric'] = FAFLmetric
        # data['FAsus'] = FAsus
        # data['FAmetric'] = FAmetric
        # FAsus = data['FAsus']
        # FAmetric = data['FAmetric']
        data['sus_nofs'] = sus
        data['metric_nofs'] = metric
        data['FAsus_nofs'] = FAsus
        data['FAmetric_nofs'] = FAmetric
        data['percent_nofs'] = percent
        data['PSsus_nofs'] = PSsus
        data['PSmetric_nofs'] = PSmetric
        data['FAFLsus_nofs'] = FAFLsus
        data['FAFLmetric_nofs'] = FAFLmetric
        data['MMsus_nofs'] = MMsus
        data['MMmetric_nofs'] = MMmetric
        # data['sus_nofs'] = sus_nofs
        # data['metric_nofs'] = metric_nofs
        # data['FAsus_nofs'] = FAsus_nofs
        # data['FAmetric_nofs'] = FAmetric_nofs
        # data['percent_nofs'] = percent_nofs
        # data = {
        #     'sus': sus, 'metric': metric, 'MGS': MGS, 'Flag': Flag,
        #     'percent': percent, 'FAsus': FAsus, 'FAmetric': FAmetric,
        #     'Exec': ExecutableS, 'Exel': ExelineS
        # }
        json_str = json.dumps(data)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'w') as f:
            json.dump(json_str, f)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # Flag = data['Flag']
        # MGS = data['MGS']
        # sus = data['sus']
        # FAsus = data['FAsus']
        # percent = data['percent']
        row = eval('getMetrics_3_test')(row, ws, mu, MGS, sus, FAsus, PSsus, FAFLsus, MMsus, sbflsus, Flag, percent, pot, pof, pal, staDe)  # +sys.argv[1][-1]
        # row = eval('getMetrics_4_test')(row, ws, mu, FAFLVariantsus, Flag)
    wb.save(path)


